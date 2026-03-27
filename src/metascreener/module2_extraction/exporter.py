"""Excel export engine.

Writes extraction results to a new Excel file with:
- One sheet per data sheet in the schema (only EXTRACT fields)
- A _MetaScreener_Log sheet with extraction metadata
"""
from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
import structlog

from metascreener.core.enums import Confidence, FieldRole
from metascreener.core.models_extraction import (
    ExtractionSchema,
    ExtractionSessionResult,
)

log = structlog.get_logger()


def export_to_excel(
    *,
    schema: ExtractionSchema,
    results: list[ExtractionSessionResult],
    output_path: Path,
    template_path: Path | None = None,
) -> Path:
    """Export extraction results to Excel.

    Args:
        schema: The ExtractionSchema used for extraction.
        results: List of per-PDF extraction results.
        output_path: Where to write the Excel file.
        template_path: Optional original template to preserve formatting.

    Returns:
        Path to the written file.
    """
    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_schema in schema.data_sheets:
        extract_fields = sheet_schema.extract_fields
        if not extract_fields:
            continue

        # Create or get sheet
        if sheet_schema.sheet_name in wb.sheetnames:
            ws = wb[sheet_schema.sheet_name]
        else:
            ws = wb.create_sheet(sheet_schema.sheet_name)

        # Write headers
        for col_idx, field in enumerate(extract_fields, start=1):
            ws.cell(row=1, column=col_idx, value=field.name)

        # Write data rows
        current_row = 2
        for pdf_result in results:
            sheet_result = pdf_result.sheets.get(sheet_schema.sheet_name)
            if not sheet_result:
                continue
            for row_result in sheet_result.rows:
                for col_idx, field in enumerate(extract_fields, start=1):
                    cell_value = row_result.fields.get(field.name)
                    if cell_value is not None:
                        ws.cell(row=current_row, column=col_idx, value=cell_value.value)
                current_row += 1

    # Add metadata log sheet
    _write_log_sheet(wb, schema, results)

    wb.save(output_path)
    log.info("excel_exported", path=str(output_path), sheets=len(schema.data_sheets),
             pdfs=len(results))
    return output_path


def _write_log_sheet(
    wb: Any,
    schema: ExtractionSchema,
    results: list[ExtractionSessionResult],
) -> None:
    """Write the _MetaScreener_Log sheet with extraction metadata."""
    if "_MetaScreener_Log" in wb.sheetnames:
        ws = wb["_MetaScreener_Log"]
    else:
        ws = wb.create_sheet("_MetaScreener_Log")

    ws.cell(row=1, column=1, value="MetaScreener Extraction Log")
    ws.cell(row=2, column=1, value="Schema ID")
    ws.cell(row=2, column=2, value=schema.schema_id)
    ws.cell(row=3, column=1, value="Schema Version")
    ws.cell(row=3, column=2, value=schema.schema_version)
    ws.cell(row=4, column=1, value="Export Date")
    ws.cell(row=4, column=2, value=datetime.now(UTC).isoformat())
    ws.cell(row=5, column=1, value="PDFs Extracted")
    ws.cell(row=5, column=2, value=len(results))

    # Per-PDF summary
    ws.cell(row=7, column=1, value="PDF")
    ws.cell(row=7, column=2, value="Sheets")
    ws.cell(row=7, column=3, value="Total Cells")
    ws.cell(row=7, column=4, value="HIGH")
    ws.cell(row=7, column=5, value="MEDIUM")
    ws.cell(row=7, column=6, value="LOW")

    row = 8
    for pdf_result in results:
        total = high = medium = low = 0
        for sr in pdf_result.sheets.values():
            for rr in sr.rows:
                for cv in rr.fields.values():
                    total += 1
                    if cv.confidence == Confidence.HIGH:
                        high += 1
                    elif cv.confidence == Confidence.MEDIUM:
                        medium += 1
                    elif cv.confidence == Confidence.LOW:
                        low += 1

        ws.cell(row=row, column=1, value=pdf_result.pdf_filename)
        ws.cell(row=row, column=2, value=len(pdf_result.sheets))
        ws.cell(row=row, column=3, value=total)
        ws.cell(row=row, column=4, value=high)
        ws.cell(row=row, column=5, value=medium)
        ws.cell(row=row, column=6, value=low)
        row += 1
