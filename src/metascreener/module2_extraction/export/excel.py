"""Enhanced Excel exporter with confidence color coding.

Exports flat cell-dict results (from the extraction repository) to an
openpyxl workbook.  Each unique pdf_id becomes one row; columns follow
the ordered field_names list.  Cells are optionally color-coded by the
confidence level of the extracted value.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import structlog
from openpyxl.styles import PatternFill

log = structlog.get_logger()

# RGB hex strings (no alpha prefix) — one per Confidence enum value.
CONFIDENCE_COLORS: dict[str, str] = {
    "VERIFIED": "15803d",  # deep green
    "HIGH":     "22c55e",  # green
    "MEDIUM":   "eab308",  # yellow
    "LOW":      "f97316",  # orange
    "SINGLE":   "a3a3a3",  # gray
    "FAILED":   "ef4444",  # red
}


def export_extraction_results(
    results: list[dict[str, str]],
    field_names: list[str],
    output_path: Path,
    include_evidence: bool = False,
    include_confidence: bool = True,
) -> Path:
    """Export extraction results to Excel with optional color coding.

    Args:
        results: Cell data from repository, each entry is a dict with keys:
            ``pdf_id``, ``field_name``, ``value``, ``confidence``,
            ``evidence_json``.
        field_names: Ordered list of field names for columns.
        output_path: Destination path for the Excel file.
        include_evidence: When True, an evidence column is appended after
            each data column.
        include_confidence: When True, cells are filled with the color that
            corresponds to their confidence level.

    Returns:
        Path to the saved file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    # ------------------------------------------------------------------ #
    # Headers
    # ------------------------------------------------------------------ #
    col = 1
    for name in field_names:
        ws.cell(1, col, name)
        if include_evidence:
            ws.cell(1, col + 1, f"{name} [evidence]")
            col += 2
        else:
            col += 1

    # ------------------------------------------------------------------ #
    # Group results by pdf_id → field_name → cell_dict
    # ------------------------------------------------------------------ #
    by_pdf: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    for cell in results:
        by_pdf[cell["pdf_id"]][cell["field_name"]] = cell

    # ------------------------------------------------------------------ #
    # Data rows — one row per pdf_id
    # ------------------------------------------------------------------ #
    row = 2
    for _pdf_id, fields in by_pdf.items():
        col = 1
        for name in field_names:
            cell_data = fields.get(name, {})
            value = cell_data.get("value", "")
            confidence = cell_data.get("confidence", "")

            ws_cell = ws.cell(row, col, value)
            if include_confidence and confidence in CONFIDENCE_COLORS:
                hex_color = CONFIDENCE_COLORS[confidence]
                ws_cell.fill = PatternFill(
                    start_color=hex_color,
                    end_color=hex_color,
                    fill_type="solid",
                )

            if include_evidence:
                evidence = cell_data.get("evidence_json", "")
                ws.cell(row, col + 1, evidence)
                col += 2
            else:
                col += 1

        row += 1

    # ------------------------------------------------------------------ #
    # Metadata sheet
    # ------------------------------------------------------------------ #
    meta = wb.create_sheet("_MetaScreener_Log")
    meta["A1"] = "Export Date"
    meta["B1"] = datetime.now(UTC).isoformat()
    meta["A2"] = "Total PDFs"
    meta["B2"] = len(by_pdf)

    wb.save(output_path)
    log.info(
        "excel_exported",
        path=str(output_path),
        pdf_count=len(by_pdf),
        fields=len(field_names),
    )
    return output_path
