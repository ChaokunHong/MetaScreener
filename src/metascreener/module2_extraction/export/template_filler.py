"""Template-based Excel export — fills the user's original template.

Instead of creating a flat export file, this module copies the original
uploaded Excel template and fills only the EXTRACT-role columns with
extracted values.  All formulas, data validations, formatting, mapping
sheets, and documentation sheets are preserved exactly as uploaded.

Multi-PDF strategy:
  - For ONE_PER_STUDY sheets: each PDF occupies exactly one Excel row,
    appended sequentially starting from row 2.
  - For MANY_PER_STUDY sheets: each PDF's rows are appended sequentially.
    A ``_pdf_filename`` column is prepended so the user knows which PDF
    each row came from.
"""
from __future__ import annotations

import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
import structlog

from metascreener.core.enums import FieldRole, SheetCardinality, SheetRole
from metascreener.core.models_extraction import ExtractionSchema

log = structlog.get_logger(__name__)

def export_filled_template(
    template_path: Path,
    results: list[dict],
    schema: ExtractionSchema,
    output_path: Path,
) -> Path:
    """Export by filling the original template with extracted values.

    Strategy:
        1. Copy the original template as-is (preserves ALL formatting,
           formulas, data validations, conditional formatting, etc.).
        2. For each data sheet, find EXTRACT-role columns via the schema.
        3. Group results by ``(pdf_id, sheet_name)`` to correctly assign
           each PDF's data to separate rows.
        4. For ONE_PER_STUDY sheets, each PDF gets exactly one row.
        5. For MANY_PER_STUDY sheets, each PDF may produce multiple rows
           which are appended sequentially.
        6. Do NOT touch formula columns, mapping sheets, or documentation
           sheets.
        7. Save as a new file at *output_path*.

    Args:
        template_path: Path to the original uploaded Excel template.
        results: Flat list of extracted cell dicts from the repository.
            Each dict has keys: ``pdf_id``, ``sheet_name``, ``field_name``,
            ``value``, and optionally ``row_index`` (int, 0-based).
        schema: Compiled ExtractionSchema with field roles and sheet roles.
        output_path: Destination path for the filled workbook.

    Returns:
        The *output_path* after successful save.

    Raises:
        FileNotFoundError: If *template_path* does not exist.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Step 1 — copy template verbatim
    shutil.copy2(template_path, output_path)

    # Step 2 — open the copy (keep formulas intact: data_only=False)
    wb = openpyxl.load_workbook(output_path)

    # Step 3 — group results by (pdf_id, sheet_name, row_index)
    # Structure: {sheet_name: {pdf_id: {row_index: {field_name: value}}}}
    grouped: dict[str, dict[str, dict[int, dict[str, str]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(dict))
    )
    # Track PDF ordering by first appearance
    pdf_order: dict[str, int] = {}
    # Also collect pdf filenames if available in results
    pdf_filenames: dict[str, str] = {}

    for cell in results:
        sheet_name = cell.get("sheet_name", "")
        pdf_id = cell.get("pdf_id", "")
        row_index = cell.get("row_index", 0)
        if not isinstance(row_index, int):
            try:
                row_index = int(row_index)
            except (ValueError, TypeError):
                row_index = 0
        field_name = cell.get("field_name", "")
        value = cell.get("value", "")

        if sheet_name and field_name and pdf_id:
            grouped[sheet_name][pdf_id][row_index][field_name] = value
            if pdf_id not in pdf_order:
                pdf_order[pdf_id] = len(pdf_order)
            # Capture filename from results if present
            if "pdf_filename" in cell and cell["pdf_filename"]:
                pdf_filenames[pdf_id] = cell["pdf_filename"]

    # Sort PDFs by first appearance
    sorted_pdf_ids = sorted(pdf_order.keys(), key=lambda pid: pdf_order[pid])

    filled_count = 0
    schema_map = {s.sheet_name: s for s in schema.sheets}

    # Step 4 — iterate data sheets only
    for sheet_schema in schema.sheets:
        if sheet_schema.role != SheetRole.DATA:
            continue

        if sheet_schema.sheet_name not in wb.sheetnames:
            log.warning(
                "sheet_not_in_workbook",
                sheet_name=sheet_schema.sheet_name,
                available=wb.sheetnames,
            )
            continue

        ws = wb[sheet_schema.sheet_name]
        sheet_name = sheet_schema.sheet_name
        is_many = sheet_schema.cardinality == SheetCardinality.MANY_PER_STUDY

        # Build header -> column index mapping (1-based)
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is not None:
                header_map[str(header).strip()] = col

        # Identify EXTRACT-role fields and their column indices
        extract_cols: dict[str, int] = {}
        for field in sheet_schema.fields:
            if field.role == FieldRole.EXTRACT:
                col_idx = header_map.get(field.name)
                if col_idx is not None:
                    extract_cols[field.name] = col_idx

        if not extract_cols:
            continue

        sheet_data = grouped.get(sheet_name, {})
        if not sheet_data:
            continue

        # Current write row (start after header)
        excel_row = 2

        for pdf_id in sorted_pdf_ids:
            pdf_rows = sheet_data.get(pdf_id)
            if not pdf_rows:
                continue

            # Sort rows by row_index
            sorted_row_indices = sorted(pdf_rows.keys())

            for row_idx in sorted_row_indices:
                fields_data = pdf_rows[row_idx]

                for field_name, col_idx in extract_cols.items():
                    value = fields_data.get(field_name)
                    if value is None or value == "":
                        continue

                    # Guard: do not overwrite formula cells
                    existing = ws.cell(excel_row, col_idx).value
                    if existing is not None and str(existing).startswith("="):
                        log.debug(
                            "skipping_formula_cell",
                            sheet=sheet_name,
                            row=excel_row,
                            col=col_idx,
                            field=field_name,
                        )
                        continue

                    ws.cell(excel_row, col_idx, value)
                    filled_count += 1

                excel_row += 1

    wb.save(output_path)
    log.info(
        "filled_template_exported",
        template=str(template_path),
        output=str(output_path),
        cells_filled=filled_count,
        n_pdfs=len(sorted_pdf_ids),
    )
    return output_path
