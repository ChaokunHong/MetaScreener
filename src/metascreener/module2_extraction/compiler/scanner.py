"""Step 1: Scan Excel template structure.

Reads an .xlsx file and produces a list of RawSheetInfo objects describing
each sheet's columns, data types, formulas, dropdowns, and sample data.
This is a pure structural scan — no semantic interpretation.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import structlog

log = structlog.get_logger()

_MAX_SAMPLE_ROWS = 10


@dataclass
class RawFieldInfo:
    """Raw information about a single column from scanning."""
    column_letter: str
    name: str
    has_formula: bool = False
    inferred_type: str = "text"
    dropdown_options: list[str] | None = None
    sample_values: list[Any] = field(default_factory=list)


@dataclass
class RawSheetInfo:
    """Raw information about a single sheet from scanning."""
    sheet_name: str
    fields: list[RawFieldInfo]
    row_count: int
    sample_row_count: int


def scan_template(path: Path) -> list[RawSheetInfo]:
    """Scan an Excel template and return raw structural info per sheet."""
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=False)
    wb_data = openpyxl.load_workbook(path, data_only=True)
    sheets: list[RawSheetInfo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        if ws.max_row is None or ws.max_row < 1:
            continue

        # Read headers from row 1
        headers: list[str] = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value is not None else "")

        if not any(headers):
            continue

        # Build dropdown map from data validations
        dropdown_map: dict[str, list[str]] = {}
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                options = [o.strip() for o in dv.formula1.strip('"').split(",")]
                for cell_range in dv.sqref.ranges:
                    for col in range(cell_range.min_col, cell_range.max_col + 1):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        dropdown_map[col_letter] = options

        # Scan fields
        fields: list[RawFieldInfo] = []
        data_row_count = max(0, (ws.max_row or 1) - 1)
        sample_rows = min(data_row_count, _MAX_SAMPLE_ROWS)

        for col_idx, header in enumerate(headers, start=1):
            if not header:
                continue

            col_letter = openpyxl.utils.get_column_letter(col_idx)

            # Check for formulas in first data row
            has_formula = False
            if ws.max_row and ws.max_row >= 2:
                cell = ws.cell(row=2, column=col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    has_formula = True

            # Collect sample values (from data-only workbook)
            sample_values: list[Any] = []
            for row_idx in range(2, 2 + sample_rows):
                val = ws_data.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    sample_values.append(val)

            # Infer type from sample values
            inferred_type = _infer_type(sample_values)

            # Check dropdown
            dropdown_options = dropdown_map.get(col_letter)

            fields.append(RawFieldInfo(
                column_letter=col_letter,
                name=header,
                has_formula=has_formula,
                inferred_type=inferred_type,
                dropdown_options=dropdown_options,
                sample_values=sample_values,
            ))

        sheets.append(RawSheetInfo(
            sheet_name=sheet_name,
            fields=fields,
            row_count=data_row_count,
            sample_row_count=sample_rows,
        ))

    wb.close()
    wb_data.close()

    log.info("template_scanned", sheets=len(sheets),
             total_fields=sum(len(s.fields) for s in sheets))
    return sheets


def _infer_type(values: list[Any]) -> str:
    """Infer field type from sample values."""
    if not values:
        return "text"

    type_counts: dict[str, int] = {"number": 0, "text": 0, "date": 0, "boolean": 0}
    for v in values:
        if isinstance(v, bool):
            type_counts["boolean"] += 1
        elif isinstance(v, (int, float)):
            type_counts["number"] += 1
        elif isinstance(v, str):
            type_counts["text"] += 1
        else:
            type_counts["text"] += 1

    return max(type_counts, key=type_counts.get)  # type: ignore[arg-type]
