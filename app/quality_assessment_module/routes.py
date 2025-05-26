from flask import render_template, request, redirect, url_for, flash, jsonify, current_app, send_from_directory, session
from . import quality_bp # Import the blueprint
from werkzeug.utils import secure_filename # For securing filenames
import os
import json # Import Python's json library
import uuid # For generating batch IDs
from typing import Optional # Added for type hint

# We will also need to import services and forms later
from .services import process_uploaded_document, quick_upload_document, get_assessment_result, QUALITY_ASSESSMENT_TOOLS, _assessments_db, QA_PDF_UPLOAD_DIR, register_celery_item, _generate_safe_assessment_id, _save_assessment_to_redis, _save_assessments_to_file
from config.config import get_api_key_for_provider, get_base_url_for_provider
# from .forms import DocumentUploadForm, AssessmentReviewForm 

# Import Redis storage utilities
from .redis_storage import save_batch_status, get_batch_status, update_batch_status, delete_batch_status

# Import Celery task for async processing
try:
    from app.celery_tasks.tasks import process_quality_assessment
    CELERY_ENABLED = True
except ImportError:
    CELERY_ENABLED = False
    process_quality_assessment = None

# Legacy in-memory storage for backward compatibility (will be phased out)
_batch_assessments_status = {}

def get_batch_info(batch_id: str):
    """Get batch info from Redis first, fallback to memory storage"""
    # Try Redis first
    batch_info = get_batch_status(batch_id)
    if batch_info:
        return batch_info
    
    # Fallback to memory storage for backward compatibility
    return _batch_assessments_status.get(batch_id)

def save_batch_info(batch_id: str, batch_data: dict):
    """Save batch info to both Redis and memory storage"""
    # Save to Redis (primary storage)
    save_batch_status(batch_id, batch_data)
    
    # Also save to memory for backward compatibility
    _batch_assessments_status[batch_id] = batch_data

# Allowed extensions for PDF files
ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Add root route for quality assessment main page
@quality_bp.route('/')
def quality_assessment_main():
    """Main quality assessment page - redirect to upload page"""
    return redirect(url_for('.upload_document_for_assessment'))

@quality_bp.route('/upload', methods=['GET', 'POST'])
def upload_document_for_assessment():
    if request.method == 'POST':
        uploaded_files = request.files.getlist("pdf_files")
        current_app.logger.info(f"BATCH_UPLOAD: Received {len(uploaded_files)} files.")
        
        if not uploaded_files or not any(f.filename for f in uploaded_files):
            flash('No PDF files selected or all files are empty.', 'error')
            current_app.logger.warning("BATCH_UPLOAD: No valid files found in submission.")
            return redirect(request.url)

        selected_document_type = request.form.get('document_type')
        successful_uploads = []
        failed_uploads = []
        assessment_ids_in_batch = []

        # Use optimized upload processing
        return handle_ultra_quick_upload(uploaded_files, selected_document_type, successful_uploads, failed_uploads, assessment_ids_in_batch)

            
    return render_template('quality_assessment_upload.html', assessment_tools_info=QUALITY_ASSESSMENT_TOOLS)

def handle_ultra_quick_upload(uploaded_files, selected_document_type, successful_uploads, failed_uploads, assessment_ids_in_batch):
    """Ultra-fast upload handler that minimizes blocking operations"""
    import tempfile
    from gevent import spawn
    
    # Pre-generate batch ID and assessment IDs
    batch_id = str(uuid.uuid4())
    temp_file_mapping = {}
    
    # Step 1: Minimal validation and create temporary references (no disk I/O yet)
    for file_storage in uploaded_files:
        if file_storage and file_storage.filename and allowed_file(file_storage.filename):
            original_filename = secure_filename(file_storage.filename)
            assessment_id = _generate_safe_assessment_id()
            
            # Store file content in memory temporarily
            file_storage.stream.seek(0)
            file_content = file_storage.stream.read()
            file_storage.stream.seek(0)
            
            # Create minimal assessment record (in-memory only)
            temp_file_mapping[assessment_id] = {
                'filename': original_filename,
                'content': file_content,
                'size': len(file_content)
            }
            
            # Create basic assessment record immediately for frontend status queries
            _assessments_db[assessment_id] = {
                "status": "uploading",
                "filename": original_filename,
                "document_type": selected_document_type,
                "progress": {"current": 0, "total": 100, "message": "File uploaded, preparing for processing"},
                "saved_pdf_filename": None,
                "raw_text": None,
                "assessment_details": [],

            }
            
            # Save to Redis for cross-process access
            try:
                _save_assessment_to_redis(assessment_id, _assessments_db[assessment_id])
            except Exception as e:
                current_app.logger.error(f"ULTRA_QUICK: Failed to save assessment {assessment_id} to Redis: {e}")
            
            successful_uploads.append(original_filename)
            assessment_ids_in_batch.append(assessment_id)
            current_app.logger.info(f"ULTRA_QUICK: Pre-processed {original_filename} (ID: {assessment_id})")
        
        elif file_storage and file_storage.filename:
            failed_uploads.append(secure_filename(file_storage.filename) + " (invalid type)")
    
    if not successful_uploads:
        return handle_upload_results(successful_uploads, failed_uploads, assessment_ids_in_batch, uploaded_files)
    
    # Save all assessment records to file after processing all files
    try:
        _save_assessments_to_file()
        current_app.logger.info("ULTRA_QUICK: Saved all assessment records to file")
    except Exception as e:
        current_app.logger.error(f"ULTRA_QUICK: Failed to save assessments to file: {e}")
    
    # Step 2: Create batch info immediately (lightweight operation)
    batch_data = {
        "status": "uploading",  # New intermediate status
        "assessment_ids": assessment_ids_in_batch,
        "total_files": len(assessment_ids_in_batch),
        "original_attempt_count": len(uploaded_files),
        "successful_filenames": successful_uploads,
        "failed_filenames": failed_uploads,

    }
    
    # Step 3: Quick Redis save (use pipeline for efficiency)
    try:
        from .redis_storage import get_redis_client
        redis_client = get_redis_client()
        pipe = redis_client.pipeline()
        pipe.setex(f"qa_batch:{batch_id}", 604800, json.dumps(batch_data))
        pipe.execute()
        current_app.logger.info(f"ULTRA_QUICK: Batch {batch_id} saved to Redis via pipeline")
    except Exception as e:
        current_app.logger.error(f"ULTRA_QUICK: Redis save failed: {e}")
        # Fallback to memory storage
        _batch_assessments_status[batch_id] = batch_data
    
    # Step 4: Spawn async task for heavy I/O operations (non-blocking)
    # Pass the current app to the async function for context
    from gevent import spawn
    spawn(process_files_async, current_app._get_current_object(), temp_file_mapping, selected_document_type, batch_id, dict(session))
    
    # Step 5: Immediate redirect (no waiting for file I/O)
    flash(f'{len(successful_uploads)} document(s) queued for ultra-fast processing. {len(failed_uploads)} failed.', 'success')
    current_app.logger.info(f"ULTRA_QUICK: Immediate redirect to batch {batch_id}")
    return redirect(url_for('.view_batch_status', batch_id=batch_id))

def process_files_async(app, temp_file_mapping, selected_document_type, batch_id, session_data):
    """Async processing of file I/O operations"""
    from gevent import sleep
    from flask import current_app
    
    # Create application context for background processing
    with app.app_context():
        current_app.logger.info(f"ASYNC_PROCESSOR: Starting background processing for batch {batch_id}")
        
        # Update batch status to processing
        try:
            update_batch_status(batch_id, {"status": "processing"})
        except Exception as e:
            current_app.logger.error(f"ASYNC_PROCESSOR: Failed to update batch status: {e}")
        
        # Process each file
        for assessment_id, file_info in temp_file_mapping.items():
            try:
                # Create file stream from content
                import io
                file_stream = io.BytesIO(file_info['content'])
                
                # Update status to processing
                if assessment_id in _assessments_db:
                    _assessments_db[assessment_id]["status"] = "processing_upload"
                    _assessments_db[assessment_id]["progress"]["message"] = "Processing file upload..."
                    _save_assessment_to_redis(assessment_id, _assessments_db[assessment_id])
                
                # Use existing assessment ID for processing
                import os
                from werkzeug.utils import secure_filename
                
                # Save PDF file
                secure_filename_for_save = secure_filename(file_info['filename'])
                saved_pdf_filename = f"{assessment_id}_{secure_filename_for_save}"
                saved_pdf_full_path = os.path.join(QA_PDF_UPLOAD_DIR, saved_pdf_filename)
                
                with open(saved_pdf_full_path, 'wb') as f_out:
                    f_out.write(file_info['content'])
                
                # Update assessment record with file info
                if assessment_id in _assessments_db:
                    _assessments_db[assessment_id]["saved_pdf_filename"] = saved_pdf_filename
                    _assessments_db[assessment_id]["status"] = "pending_text_extraction"
                    _assessments_db[assessment_id]["progress"]["message"] = "File saved, preparing for assessment"
                    _save_assessment_to_redis(assessment_id, _assessments_db[assessment_id])
                    # Save to file for persistence
                    try:
                        from app.quality_assessment_module.services import _save_assessments_to_file
                        _save_assessments_to_file()
                    except Exception as save_e:
                        current_app.logger.error(f"ASYNC_PROCESSOR: Failed to save to file: {save_e}")
                
                # Schedule background processing with existing ID
                from app.quality_assessment_module.services import run_background_processing
                from config.config import get_current_llm_config
                from gevent import spawn
                
                # Get LLM config and ensure it's valid
                llm_config = get_current_llm_config(session_data)
                
                # Ensure all required fields are present
                if llm_config:
                    provider_name = llm_config.get('provider_name')
                    api_key_val = get_api_key_for_provider(provider_name, session_data)
                    base_url = get_base_url_for_provider(provider_name)
                    
                    complete_llm_config = {
                        "provider_name": provider_name,
                        "model_id": llm_config.get('model_id'),
                        "base_url": base_url,
                        "api_key": api_key_val
                    }
                    
                    current_app.logger.info(f"ASYNC_PROCESSOR: Starting background processing for {assessment_id} with complete LLM config")
                    spawn(run_background_processing, assessment_id, current_app.app_context(), complete_llm_config)
                else:
                    current_app.logger.error(f"ASYNC_PROCESSOR: No LLM config available for {assessment_id}")
                    # Update assessment status to error
                    if assessment_id in _assessments_db:
                        _assessments_db[assessment_id]["status"] = "error"
                        _assessments_db[assessment_id]["message"] = "LLM configuration not available"
                        _assessments_db[assessment_id]["progress"]["message"] = "Configuration error"
                        _save_assessment_to_redis(assessment_id, _assessments_db[assessment_id])
                
                current_app.logger.info(f"ASYNC_PROCESSOR: Completed processing {file_info['filename']} (ID: {assessment_id})")
                
                # Small delay to prevent overwhelming the system
                sleep(0.1)
                
            except Exception as e:
                current_app.logger.error(f"ASYNC_PROCESSOR: Error processing {file_info['filename']}: {e}")
                # Update assessment record with error status
                if assessment_id in _assessments_db:
                    _assessments_db[assessment_id]["status"] = "error"
                    _assessments_db[assessment_id]["message"] = f"Async processing failed: {str(e)}"
                    _assessments_db[assessment_id]["progress"]["message"] = "Processing failed"
                    _save_assessment_to_redis(assessment_id, _assessments_db[assessment_id])
        
        current_app.logger.info(f"ASYNC_PROCESSOR: Completed background processing for batch {batch_id}")

def handle_upload_results(successful_uploads, failed_uploads, assessment_ids_in_batch, uploaded_files):
    """Handle upload results and determine redirect behavior"""
    if successful_uploads:
        if len(successful_uploads) == 1 and not failed_uploads:
            flash(f'Document \'{successful_uploads[0]}\' uploaded. Assessment process initiated.', 'success')
            current_app.logger.info(f"BATCH_UPLOAD: Single file success ({successful_uploads[0]}), redirecting to individual result.")
            return redirect(url_for('.view_assessment_result', assessment_id=assessment_ids_in_batch[0]))
        else:
            batch_id = str(uuid.uuid4())
            batch_data = {
                "status": "processing",
                "assessment_ids": assessment_ids_in_batch,
                "total_files": len(assessment_ids_in_batch),
                "original_attempt_count": len(uploaded_files),
                "successful_filenames": successful_uploads,
                "failed_filenames": failed_uploads
            }
            # Save to Redis instead of memory
            save_batch_info(batch_id, batch_data)
            flash(f'{len(successful_uploads)} document(s) queued for assessment. {len(failed_uploads)} failed.', 'info')
            current_app.logger.info(f"BATCH_UPLOAD: Multiple files ({len(successful_uploads)} success, {len(failed_uploads)} failed). Batch ID: {batch_id}. Redirecting to batch status.")
            current_app.logger.info(f"BATCH_UPLOAD: Batch data for {batch_id}: {batch_data}")
            return redirect(url_for('.view_batch_status', batch_id=batch_id))

    elif failed_uploads:
        flash(f'All file uploads failed. Reasons: {", ".join(failed_uploads)}', 'error')
        current_app.logger.warning(f"BATCH_UPLOAD: All file uploads failed.")
        return redirect(request.url)
    else: 
        flash('No files were processed.', 'warning')
        current_app.logger.warning("BATCH_UPLOAD: No files were processed (edge case).")
        return redirect(request.url)

@quality_bp.route('/async_upload', methods=['POST'])
def async_upload_documents():
    """Asynchronous quality assessment upload using Celery"""
    if not CELERY_ENABLED:
        return jsonify({
            'status': 'error',
            'message': 'Async processing not available. Celery is not configured.'
        }), 503
    
    temp_files_info_for_celery = [] # Initialize here to be in scope for finally/except cleanup
    try:
        uploaded_files = request.files.getlist("pdf_files")
        current_app.logger.info(f"ASYNC_QA_UPLOAD: Received {len(uploaded_files)} files in request.")
        
        if not uploaded_files or not any(f.filename for f in uploaded_files):
            return jsonify({'status': 'error', 'message': 'No PDF files uploaded'}), 400
        
        for uploaded_file in uploaded_files:
            if uploaded_file.filename != '' and allowed_file(uploaded_file.filename):
                secure_name = secure_filename(uploaded_file.filename)
                temp_file_path = os.path.join(QA_PDF_UPLOAD_DIR, f"celery_async_qa_{uuid.uuid4()}_{secure_name}")
                uploaded_file.save(temp_file_path)
                temp_files_info_for_celery.append({
                    'path': temp_file_path,      
                    'original_filename': uploaded_file.filename 
                })
        
        if not temp_files_info_for_celery:
            return jsonify({'status': 'error', 'message': 'No valid PDF files were processed for async task'}), 400
        
        selected_document_type = request.form.get('document_type')
        assessment_config = {
            'document_type': selected_document_type,
            'tools_info': QUALITY_ASSESSMENT_TOOLS 
        }
        
        from config.config import get_current_llm_config
        session_dict = dict(session)  # Convert session to dict early for safety
        llm_config = get_current_llm_config(session)
        
        celery_processing_uuid = str(uuid.uuid4()) 
        frontend_batch_uuid = str(uuid.uuid4())
        
        numerical_item_ids = []
        for file_info_for_celery in temp_files_info_for_celery:
            original_fname = file_info_for_celery['original_filename']
            try:
                num_id = register_celery_item(
                    filename=original_fname,
                    document_type=assessment_config['document_type'],
                    celery_processing_uuid=celery_processing_uuid
                )
                numerical_item_ids.append(num_id)
                current_app.logger.info(f"ASYNC_QA_UPLOAD: Registered item '{original_fname}' with numerical ID {num_id} for Celery batch {frontend_batch_uuid} (Celery proc. UUID: {celery_processing_uuid})")
                
                # Immediately save to Redis for multi-process access
                from app.quality_assessment_module.services import _assessments_db, _save_assessment_to_redis
                if num_id in _assessments_db:
                    _save_assessment_to_redis(num_id, _assessments_db[num_id])
                    current_app.logger.info(f"ASYNC_QA_UPLOAD: Saved {num_id} to Redis for immediate cross-process access")
                else:
                    current_app.logger.warning(f"ASYNC_QA_UPLOAD: {num_id} not found in _assessments_db after registration")
                    
            except Exception as e_reg:
                current_app.logger.error(f"ASYNC_QA_UPLOAD: Failed to register item '{original_fname}' for Celery: {e_reg}")
        
        if not numerical_item_ids:
             current_app.logger.error(f"ASYNC_QA_UPLOAD: No items were successfully registered for Celery processing.")
             # Clean up already saved temp files if registration fails for all
             for finfo_cleanup in temp_files_info_for_celery:
                 if 'path' in finfo_cleanup and os.path.exists(finfo_cleanup['path']):
                    try: os.remove(finfo_cleanup['path'])
                    except: pass
             return jsonify({'status': 'error', 'message': 'Failed to prepare any items for asynchronous assessment.'}), 500

        batch_info_for_redis = {
            "assessment_ids": numerical_item_ids, 
            "status": "processing_celery",        
            "celery_processing_uuid": celery_processing_uuid, 
            "total_files_in_batch": len(numerical_item_ids),
            "original_upload_count": len(uploaded_files)
        }
        save_batch_status(frontend_batch_uuid, batch_info_for_redis) 
        current_app.logger.info(f"ASYNC_QA_UPLOAD: Saved batch info for {frontend_batch_uuid} to Redis (qa_batch:{frontend_batch_uuid}) with items: {numerical_item_ids}")

        task = process_quality_assessment.delay(
            temp_files_info_for_celery, 
            assessment_config,
            llm_config,
            session_dict,  # Use the converted session dict
            celery_processing_uuid 
        )
        
        current_app.logger.info(f"ASYNC_QA_UPLOAD: Started Celery task {task.id} for internal processing UUID {celery_processing_uuid} (Frontend batch UUID: {frontend_batch_uuid}) with {len(temp_files_info_for_celery)} files.")
        
        return jsonify({
            'status': 'success',
            'task_id': task.id, 
            'assessment_id': frontend_batch_uuid, 
            'total_files': len(numerical_item_ids), 
            'message': 'Quality assessment (async) started. Batch ID assigned.'
        })
        
    except Exception as e:
        current_app.logger.exception(f"Error starting async quality assessment: {e}")
        for finfo in temp_files_info_for_celery: 
            if 'path' in finfo and os.path.exists(finfo['path']):
                try:
                    os.remove(finfo['path'])
                except OSError as e_remove_final:
                    current_app.logger.error(f"ASYNC_QA_UPLOAD: Error cleaning up temp file {finfo['path']} on global exception: {e_remove_final}")
                except Exception as e_clean_final_generic:
                    current_app.logger.error(f"ASYNC_QA_UPLOAD: Generic error cleaning temp file {finfo['path']} on global exception: {e_clean_final_generic}")
        return jsonify({'status': 'error', 'message': f'Server error: {str(e)}'}), 500

@quality_bp.route('/batch_status/<batch_id>')
def view_batch_status(batch_id):
    current_app.logger.info(f"BATCH_STATUS_VIEW: Accessed for batch_id: {batch_id}")
    batch_info = get_batch_info(batch_id)
    if not batch_info:
        flash("Batch assessment not found or has expired (server may have restarted).", 'error')
        current_app.logger.warning(f"BATCH_STATUS_VIEW: Batch ID {batch_id} not found in storage.")
        return redirect(url_for('.upload_document_for_assessment'))
    
    current_app.logger.info(f"BATCH_STATUS_VIEW: Rendering batch status for {batch_id} with info: {batch_info}")
    return render_template('quality_assessment_batch_status.html', batch_info=batch_info, batch_id=batch_id)

@quality_bp.route('/result/<assessment_id>')
def view_assessment_result(assessment_id):
    result_data = get_assessment_result(assessment_id)
    if not result_data:
        flash('Assessment result not found.', 'error')
        return redirect(url_for('.upload_document_for_assessment'))

    result_data_json_str = "{}" # Default to empty JSON object string
    try:
        # For initial data for JS, be very selective to avoid JSON parsing issues.
        # The polling mechanism will fetch more complete/raw data from the API.
        # Jinja server-side rendering will use the full 'result_data' object.
        selective_initial_data = {
            "status": result_data.get("status"),
            "filename": result_data.get("filename"),
            "document_type": result_data.get("document_type"),
            "progress": result_data.get("progress"),
            # DO NOT include assessment_details or raw_text here for JS init.
        }
        result_data_json_str = json.dumps(selective_initial_data)
    except Exception as e:
        current_app.logger.error(f"Error preparing result_data_json_str for JS for assessment {assessment_id}: {e}")
        # result_data_json_str remains "{}"

    current_app.logger.info(f"VIEW_RESULT_ROUTE: For assessment_id {assessment_id}, result_data being passed to template:")
    current_app.logger.info(f"VIEW_RESULT_ROUTE: result_data.status = {result_data.get('status')}")
    current_app.logger.info(f"VIEW_RESULT_ROUTE: result_data.assessment_details type = {type(result_data.get('assessment_details'))}")
    current_app.logger.info(f"VIEW_RESULT_ROUTE: result_data.assessment_details content = {result_data.get('assessment_details')}")

    return render_template(
        'quality_assessment_result.html', 
        result=result_data, # Full data for Jinja server-side rendering
        result_json_for_js=result_data_json_str, # Sanitized, minimal JSON string for JS
        assessment_id=assessment_id,
        QUALITY_ASSESSMENT_TOOLS=QUALITY_ASSESSMENT_TOOLS
    )

# --- NEW API ENDPOINT FOR ASSESSMENT STATUS --- #
@quality_bp.route('/assessment_status/<assessment_id>')
def assessment_status(assessment_id):
    current_app.logger.info(f"ASSESSMENT_STATUS: Requested ID={assessment_id}, Type={type(assessment_id)}")
    
    # Try multiple ways to get data
    assessment_data = None
    
    # 1. First try original ID
    assessment_data = get_assessment_result(assessment_id)
    current_app.logger.info(f"ASSESSMENT_STATUS: Direct lookup result for {assessment_id}: {assessment_data is not None}")
    
    # 2. If not found, try string type
    if not assessment_data:
        str_id = str(assessment_id)
        assessment_data = get_assessment_result(str_id)
        current_app.logger.info(f"ASSESSMENT_STATUS: String lookup result for {str_id}: {assessment_data is not None}")
    
    # 3. If still not found, log detailed debug information
    if not assessment_data:
        # Log current assessment data keys
        from app.quality_assessment_module.services import _assessments_db
        current_keys = list(_assessments_db.keys())
        current_app.logger.warning(f"ASSESSMENT_STATUS: No data found for {assessment_id}. Available keys: {current_keys}")
        
        # Try querying Redis directly
        try:
            from app.quality_assessment_module.services import get_assessment_redis_client
            redis_client = get_assessment_redis_client()
            redis_keys = [k.decode('utf-8') if isinstance(k, bytes) else k for k in redis_client.keys("qa_assessment:*")]
            current_app.logger.warning(f"ASSESSMENT_STATUS: Redis keys found: {redis_keys}")
        except Exception as e:
            current_app.logger.error(f"ASSESSMENT_STATUS: Redis check failed: {e}")
    
    # Return result
    if assessment_data:
        response_data = {
            "status": assessment_data.get("status"),
            "progress": assessment_data.get("progress"),
            "filename": assessment_data.get("filename"),
            "document_type": assessment_data.get("document_type")
        }
        if assessment_data.get("status") == "completed":
            pass # Reload handles showing details
        elif assessment_data.get("status") == "error":
             response_data["message"] = assessment_data.get("message", "An unknown error occurred.")

        current_app.logger.info(f"ASSESSMENT_STATUS: Returning data for {assessment_id}: status={response_data.get('status')}")
        return jsonify(response_data)
    
    current_app.logger.error(f"ASSESSMENT_STATUS: 404 - Assessment {assessment_id} not found")
    return jsonify({"status": "error", "message": "Assessment not found"}), 404
# --- END NEW API ENDPOINT --- #

# --- Debug endpoint for assessment data --- #
@quality_bp.route('/debug_assessment/<assessment_id>')
def debug_assessment_data(assessment_id):
    assessment_data = get_assessment_result(assessment_id)
    if assessment_data:
        details = assessment_data.get('assessment_details')
        details_count = len(details) if isinstance(details, list) else 0
        first_result_summary = {}
        if details and isinstance(details, list) and details_count > 0:
            first_item = details[0]
            if isinstance(first_item, dict):
                first_result_summary = {
                    "criterion_id": first_item.get("criterion_id"),
                    "judgment": first_item.get("judgment")
                }

        debug_data = {
            "status": assessment_data.get("status"),
            "document_type": assessment_data.get("document_type"),
            "has_assessment_details": bool(details),
            "assessment_details_count": details_count,
            "first_result_summary": first_result_summary, 
            "progress": assessment_data.get("progress")
        }
        return jsonify(debug_data)
    return jsonify({"error": "Assessment not found"}), 404

# --- HTML debug view for assessment details ---
@quality_bp.route('/view_details/<assessment_id>')
def view_details(assessment_id):
    """Simple HTML view for assessment details without full template complexity"""
    result_data = get_assessment_result(assessment_id)
    if not result_data:
        return f"<h1>Assessment {assessment_id} not found</h1>"
    
    html_output = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Assessment Details #{assessment_id}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .header {{ background-color: #4CAF50; color: white; padding: 15px; }}
            .container {{ margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Assessment Details: {result_data.get('filename', 'Unknown')}</h1>
            <p>Status: {result_data.get('status', 'Unknown')} | Type: {result_data.get('document_type', 'Unknown')}</p>
        </div>
        <div class="container">
    """
    
    assessment_details_list = result_data.get('assessment_details')
    if isinstance(assessment_details_list, list) and assessment_details_list:
        html_output += f"""
            <h2>Assessment Results ({len(assessment_details_list)} items)</h2>
            <table>
                <tr>
                    <th>Criterion</th>
                    <th>Judgment</th>
                    <th>Reason</th>
                </tr>
        """
        
        for item in assessment_details_list:
            html_output += f"""
                <tr>
                    <td>{item.get('criterion_text', 'No text')}</td>
                    <td>{item.get('judgment', 'No judgment')}</td>
                    <td>{item.get('reason', 'No reason')}</td>
                </tr>
            """
        
        html_output += "</table>"
    else:
        html_output += """
            <div style="color: red; padding: 20px; background-color: #ffeeee; border: 1px solid #ffcccc;">
                <h2>No assessment details found</h2>
                <p>This assessment does not contain any evaluation details.</p>
                <p>Details type: {type(assessment_details_list).__name__}, Details content: {assessment_details_list}</p>
            </div>
        """
    
    html_output += "</div></body></html>"
    
    return html_output

# --- NEW ROUTE TO SERVE SAVED PDFs FOR PREVIEW --- #
@quality_bp.route('/serve_pdf/<assessment_id>')
def serve_saved_pdf(assessment_id):
    assessment_data = get_assessment_result(assessment_id)
    if not assessment_data:
        return "Assessment data not found", 404
    
    saved_filename = assessment_data.get('saved_pdf_filename')
    if not saved_filename:
        # This case can happen if PDF saving failed during upload, or for older assessments
        # before this feature was added.
        # You could serve a placeholder PDF or an error message.
        # For now, let's return a 404 and the JS client can handle it.
        current_app.logger.warning(f"No saved_pdf_filename found for assessment_id {assessment_id}. Cannot serve PDF.")
        return "PDF record not found for this assessment or PDF was not saved.", 404

    # QA_PDF_UPLOAD_DIR is imported from services.py
    # Ensure the directory path is absolute for send_from_directory if it's not already
    # However, QA_PDF_UPLOAD_DIR should be absolute as defined in services.py
    current_app.logger.info(f"Attempting to serve PDF: {saved_filename} from directory: {QA_PDF_UPLOAD_DIR} for assessment {assessment_id}")
    try:
        return send_from_directory(QA_PDF_UPLOAD_DIR, saved_filename, as_attachment=False)
    except FileNotFoundError:
        current_app.logger.error(f"File not found: {saved_filename} in directory {QA_PDF_UPLOAD_DIR}")
        return "PDF file not found on server.", 404
    except Exception as e:
        current_app.logger.error(f"Error serving PDF {saved_filename} for assessment {assessment_id}: {e}")
        return "Error serving PDF.", 500

# --- NEW API for Batch Status --- #
@quality_bp.route('/batch_info_api/<batch_id>')
def get_batch_info_api(batch_id):
    """API endpoint for batch status information"""
    batch_info = get_batch_info(batch_id)
    if not batch_info:
        return jsonify({"error": "Batch not found"}), 404
    
    return jsonify(batch_info)

# --- NEW API for Batch Summary --- #
@quality_bp.route('/batch_summary/<batch_id>')
def get_batch_summary(batch_id):
    batch_info = get_batch_info(batch_id)
    if not batch_info:
        return jsonify({"error": "Batch not found"}), 404

    summaries = []
    assessment_ids = batch_info.get("assessment_ids", [])
    
    for idx, assessment_id in enumerate(assessment_ids):
        result_data = get_assessment_result(assessment_id)
        if result_data and result_data.get("status") == "completed":
            doc_type = result_data.get("document_type", "Unknown")
            tool_name = QUALITY_ASSESSMENT_TOOLS.get(doc_type, {}).get("tool_name", "N/A") 
            filename = result_data.get("filename", batch_info.get("successful_filenames", [])[idx] if idx < len(batch_info.get("successful_filenames", [])) else "Unknown Filename")
            
            summary_item = {
                "assessment_id": assessment_id,
                "filename": filename,
                "document_type": doc_type,
                "tool_name": tool_name,
                "total_criteria": result_data.get("summary_total_criteria_evaluated", 0),
                "negative_findings": result_data.get("summary_negative_findings", 0)
            }
            summaries.append(summary_item)
        elif result_data: # Still processing or error for this item
             summaries.append({
                "assessment_id": assessment_id,
                "filename": batch_info.get("successful_filenames", [])[idx] if idx < len(batch_info.get("successful_filenames", [])) else "Unknown Filename",
                "status": result_data.get("status", "Unknown")
            })
        # If result_data is None (shouldn't happen if it was in assessment_ids), it will be skipped

    return jsonify(summaries)

# We need to access _assessments_db for the flash message, either pass it or check status differently
# For simplicity, let's import it here. In a real app, this would be a proper database call.

# --- NEW: Quality Assessment History Route --- #
@quality_bp.route('/history')
def quality_assessment_history():
    """View history of all quality assessments and batches from the last 24 hours"""
    view_type = request.args.get('view', 'individual')  # 'individual' or 'batch'
    
    if view_type == 'batch':
        return quality_assessment_batch_history()
    else:
        return quality_assessment_individual_history()

def quality_assessment_individual_history():
    """View history of all quality assessments from the last 24 hours"""
    current_app.logger.info("QA_HISTORY: Accessing quality assessment history page")
    
    # Get all assessments from the last 24 hours
    from app.quality_assessment_module.services import _assessments_db, get_assessment_redis_client
    import time
    
    history_records = []
    current_time = time.time()
    twenty_four_hours_ago = current_time - (24 * 60 * 60)  # 24 hours in seconds
    
    # Get from memory storage first
    for assessment_id, assessment_data in _assessments_db.items():
        created_time = assessment_data.get('created_at', 0)
        
        # If no created_at timestamp, include all recent assessments (assume they're recent)
        # or use current time as fallback for very recent assessments
        if created_time == 0:
            # For assessments without timestamp, check if they have recent activity
            status = assessment_data.get('status', 'Unknown')
            if status in ['pending_text_extraction', 'processing_assessment', 'uploading', 'error']:
                created_time = current_time  # Treat as current for recent activity
            else:
                created_time = current_time - (12 * 60 * 60)  # Assume 12 hours ago for completed ones
        
        if created_time >= twenty_four_hours_ago:
            history_records.append({
                'assessment_id': assessment_id,
                'filename': assessment_data.get('filename', 'Unknown'),
                'document_type': assessment_data.get('document_type', 'Unknown'),
                'status': assessment_data.get('status', 'Unknown'),
                'created_at': created_time,
                'total_criteria': assessment_data.get('summary_total_criteria_evaluated', 0),
                'negative_findings': assessment_data.get('summary_negative_findings', 0),
                'source': 'memory'
            })
    
    # Also check Redis for any additional records
    try:
        redis_client = get_assessment_redis_client()
        redis_keys = redis_client.keys("qa_assessment:*")
        
        for key in redis_keys:
            if isinstance(key, bytes):
                key = key.decode('utf-8')
            assessment_id = key.replace("qa_assessment:", "")
            
            # Skip if already found in memory
            if assessment_id in _assessments_db:
                continue
                
            try:
                import pickle
                serialized_data = redis_client.get(key)
                if serialized_data:
                    assessment_data = pickle.loads(serialized_data)
                    created_time = assessment_data.get('created_at', 0)
                    
                    # If no created_at timestamp, include recent assessments
                    if created_time == 0:
                        status = assessment_data.get('status', 'Unknown')
                        if status in ['pending_text_extraction', 'processing_assessment', 'uploading', 'error']:
                            created_time = current_time  # Treat as current for recent activity
                        else:
                            created_time = current_time - (12 * 60 * 60)  # Assume 12 hours ago for completed ones
                    
                    if created_time >= twenty_four_hours_ago:
                        history_records.append({
                            'assessment_id': assessment_id,
                            'filename': assessment_data.get('filename', 'Unknown'),
                            'document_type': assessment_data.get('document_type', 'Unknown'),
                            'status': assessment_data.get('status', 'Unknown'),
                            'created_at': created_time,
                            'total_criteria': assessment_data.get('summary_total_criteria_evaluated', 0),
                            'negative_findings': assessment_data.get('summary_negative_findings', 0),
                            'source': 'redis'
                        })
            except Exception as e:
                current_app.logger.warning(f"QA_HISTORY: Error processing Redis key {key}: {e}")
                continue
                
    except Exception as e:
        current_app.logger.error(f"QA_HISTORY: Error accessing Redis: {e}")
    
    # Sort by creation time (newest first)
    history_records.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Add formatted time and quality score
    for record in history_records:
        record['created_at_formatted'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record['created_at']))
        if record['status'] == 'completed' and record['total_criteria'] > 0:
            record['quality_score'] = round((record['total_criteria'] - record['negative_findings']) / record['total_criteria'] * 100, 1)
        else:
            record['quality_score'] = None
    
    current_app.logger.info(f"QA_HISTORY: Found {len(history_records)} records from last 24 hours")
    
    return render_template('quality_assessment_history.html', 
                         history_records=history_records,
                         view_type='individual',
                         QUALITY_ASSESSMENT_TOOLS=QUALITY_ASSESSMENT_TOOLS)

def quality_assessment_batch_history():
    """View history of batch assessments from the last 24 hours"""
    current_app.logger.info("QA_BATCH_HISTORY: Accessing batch history page")
    
    # Get batch records from Redis
    from app.quality_assessment_module.redis_storage import get_redis_client
    import time
    
    batch_records = []
    current_time = time.time()
    twenty_four_hours_ago = current_time - (24 * 60 * 60)
    
    try:
        redis_client = get_redis_client()
        batch_keys = redis_client.keys("qa_batch:*")
        
        for key in batch_keys:
            if isinstance(key, bytes):
                key = key.decode('utf-8')
            batch_id = key.replace("qa_batch:", "")
            
            try:
                import json
                batch_data_str = redis_client.get(key)
                if batch_data_str:
                    if isinstance(batch_data_str, bytes):
                        batch_data_str = batch_data_str.decode('utf-8')
                    batch_data = json.loads(batch_data_str)
                    
                    # Get creation time (use current time if not available)
                    created_time = batch_data.get('created_at', current_time - (2 * 60 * 60))  # Default to 2 hours ago
                    
                    if created_time >= twenty_four_hours_ago:
                        # Get assessment details for this batch
                        assessment_ids = batch_data.get('assessment_ids', [])
                        completed_count = 0
                        processing_count = 0
                        error_count = 0
                        total_quality_score = 0
                        quality_scores = []
                        
                        for assessment_id in assessment_ids:
                            result_data = get_assessment_result(assessment_id)
                            if result_data:
                                status = result_data.get('status', 'unknown')
                                if status == 'completed':
                                    completed_count += 1
                                    total_criteria = result_data.get('summary_total_criteria_evaluated', 0)
                                    negative_findings = result_data.get('summary_negative_findings', 0)
                                    if total_criteria > 0:
                                        score = round((total_criteria - negative_findings) / total_criteria * 100, 1)
                                        quality_scores.append(score)
                                elif status in ['processing_assessment', 'pending_assessment', 'pending_text_extraction']:
                                    processing_count += 1
                                elif status == 'error':
                                    error_count += 1
                        
                        avg_quality_score = round(sum(quality_scores) / len(quality_scores), 1) if quality_scores else None
                        
                        batch_records.append({
                            'batch_id': batch_id,
                            'total_files': len(assessment_ids),
                            'completed_count': completed_count,
                            'processing_count': processing_count,
                            'error_count': error_count,
                            'avg_quality_score': avg_quality_score,
                            'created_at': created_time,
                            'created_at_formatted': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(created_time)),
                            'status': batch_data.get('status', 'unknown'),
                            'successful_filenames': batch_data.get('successful_filenames', [])
                        })
                        
            except Exception as e:
                current_app.logger.warning(f"QA_BATCH_HISTORY: Error processing batch {batch_id}: {e}")
                continue
                
    except Exception as e:
        current_app.logger.error(f"QA_BATCH_HISTORY: Error accessing Redis: {e}")
    
    # Sort by creation time (newest first)
    batch_records.sort(key=lambda x: x['created_at'], reverse=True)
    
    current_app.logger.info(f"QA_BATCH_HISTORY: Found {len(batch_records)} batch records from last 24 hours")
    
    return render_template('quality_assessment_history.html', 
                         history_records=batch_records,
                         view_type='batch',
                         QUALITY_ASSESSMENT_TOOLS=QUALITY_ASSESSMENT_TOOLS)

# We will need more routes:
# - Route for batch upload status / results table
# - API endpoint for AI to classify document type (if we go that route)
# - API endpoint for AI to assess based on criteria
# - Route for user to submit their own review/override AI 

# --- NEW: Enhanced Batch Summary/Results Page --- #
@quality_bp.route('/batch_results/<batch_id>')
def view_batch_results(batch_id):
    """Enhanced batch results page with comprehensive summary and export options"""
    current_app.logger.info(f"BATCH_RESULTS_VIEW: Accessed for batch_id: {batch_id}")
    batch_info = get_batch_info(batch_id)
    if not batch_info:
        flash("Batch assessment not found or has expired (server may have restarted).", 'error')
        current_app.logger.warning(f"BATCH_RESULTS_VIEW: Batch ID {batch_id} not found in storage.")
        return redirect(url_for('.upload_document_for_assessment'))
    
    # Get detailed results for all assessments in this batch
    assessment_ids = batch_info.get("assessment_ids", [])
    detailed_results = []
    results_by_type = {}
    overall_stats = {
        'total_files': len(assessment_ids),
        'completed': 0,
        'processing': 0,
        'error': 0,
        'total_criteria_evaluated': 0,
        'total_negative_findings': 0
    }
    
    for idx, assessment_id in enumerate(assessment_ids):
        result_data = get_assessment_result(assessment_id)
        if result_data:
            status = result_data.get("status", "unknown")
            doc_type = result_data.get("document_type", "Unknown")
            filename = result_data.get("filename", batch_info.get("successful_filenames", [])[idx] if idx < len(batch_info.get("successful_filenames", [])) else "Unknown Filename")
            
            # Initialize document type group if needed
            if doc_type not in results_by_type:
                results_by_type[doc_type] = {
                    'tool_name': QUALITY_ASSESSMENT_TOOLS.get(doc_type, {}).get("tool_name", "Standard Assessment"),
                    'files': [],
                    'completed_count': 0,
                    'total_criteria': 0,
                    'total_negative': 0
                }
            
            file_result = {
                "assessment_id": assessment_id,
                "filename": filename,
                "status": status,
                "document_type": doc_type
            }
            
            # Update overall stats
            if status == 'completed':
                overall_stats['completed'] += 1
                total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
                negative_findings = result_data.get("summary_negative_findings", 0)
                
                overall_stats['total_criteria_evaluated'] += total_criteria
                overall_stats['total_negative_findings'] += negative_findings
                
                # Add completed-specific data
                file_result.update({
                    "total_criteria": total_criteria,
                    "negative_findings": negative_findings,
                    "quality_score": round((total_criteria - negative_findings) / total_criteria * 100, 1) if total_criteria > 0 else 0
                })
                
                # Update type-specific stats
                results_by_type[doc_type]['completed_count'] += 1
                results_by_type[doc_type]['total_criteria'] += total_criteria
                results_by_type[doc_type]['total_negative'] += negative_findings
                
            elif status in ['processing_assessment', 'pending_assessment']:
                overall_stats['processing'] += 1
            else:
                overall_stats['error'] += 1
                if 'message' in result_data:
                    file_result['error_message'] = result_data['message']
            
            results_by_type[doc_type]['files'].append(file_result)
            detailed_results.append(file_result)
    
    # Calculate type-specific averages
    for doc_type, type_data in results_by_type.items():
        if type_data['completed_count'] > 0:
            type_data['avg_quality_score'] = round((type_data['total_criteria'] - type_data['total_negative']) / type_data['total_criteria'] * 100, 1) if type_data['total_criteria'] > 0 else 0
        else:
            type_data['avg_quality_score'] = 0
    
    # Calculate overall quality score
    overall_stats['overall_quality_score'] = round((overall_stats['total_criteria_evaluated'] - overall_stats['total_negative_findings']) / overall_stats['total_criteria_evaluated'] * 100, 1) if overall_stats['total_criteria_evaluated'] > 0 else 0
    
    current_app.logger.info(f"BATCH_RESULTS_VIEW: Rendering enhanced batch results for {batch_id}")
    return render_template('quality_assessment_batch_results.html', 
                           batch_info=batch_info, 
                           batch_id=batch_id,
                           detailed_results=detailed_results,
                           results_by_type=results_by_type,
                           overall_stats=overall_stats,
                           QUALITY_ASSESSMENT_TOOLS=QUALITY_ASSESSMENT_TOOLS)

# --- NEW: Download Individual Assessment Package (Data + Report) --- #
@quality_bp.route('/download_report/<assessment_id>/<format>')
def download_assessment_package(assessment_id, format):
    """Download assessment package: data file (Excel/CSV) + beautiful PDF report"""
    import pandas as pd
    import io
    import zipfile
    import tempfile
    import os
    from flask import send_file
    from datetime import datetime
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    
    # Get assessment data
    result_data = get_assessment_result(assessment_id)
    if not result_data or result_data.get("status") != "completed":
        flash("Assessment report not available.", 'error')
        return redirect(url_for('.quality_assessment_history'))
    
    try:
        # Prepare basic information
        filename = result_data.get('filename', 'Unknown')
        doc_type = result_data.get('document_type', 'Unknown')
        tool_name = QUALITY_ASSESSMENT_TOOLS.get(doc_type, {}).get('tool_name', 'Standard Assessment')
        
        # Calculate summary
        total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
        negative_findings = result_data.get("summary_negative_findings", 0)
        quality_score = round((total_criteria - negative_findings) / total_criteria * 100, 1) if total_criteria > 0 else 0
        
        # Generate safe filename
        safe_filename = filename.replace('.pdf', '').replace(' ', '_').replace('/', '_')
        
        # Create temporary directory for package files
        with tempfile.TemporaryDirectory() as temp_dir:
            # 1. Generate beautiful PDF report
            pdf_filename = f"assessment_report_{safe_filename}_{assessment_id}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            create_beautiful_pdf_report(pdf_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score)
            
            # 2. Generate data file (CSV or Excel)
            if format == 'csv':
                data_filename = f"assessment_data_{safe_filename}_{assessment_id}.csv"
                data_path = os.path.join(temp_dir, data_filename)
                create_csv_data_file(data_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score)
            else:  # xlsx
                data_filename = f"assessment_data_{safe_filename}_{assessment_id}.xlsx"
                data_path = os.path.join(temp_dir, data_filename)
                create_excel_data_file(data_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score)
            
            # 3. Create ZIP package
            zip_filename = f"assessment_package_{safe_filename}_{assessment_id}.zip"
            zip_path = os.path.join(temp_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(pdf_path, pdf_filename)
                zipf.write(data_path, data_filename)
                
                # Add README file
                readme_content = f"""AI-Powered Literature Quality Assessment Package
===============================================

This package contains:
1. {pdf_filename} - Professional assessment report with quality interpretation and recommendations
2. {data_filename} - Complete assessment data with detailed criteria evaluation

ASSESSMENT SUMMARY:
- Document: {filename}
- Assessment ID: {assessment_id}
- Quality Score: {quality_score}%
- Quality Rating: {'Excellent' if quality_score >= 80 else 'Good' if quality_score >= 70 else 'Fair' if quality_score >= 60 else 'Poor'}
- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

QUALITY SCORE INTERPRETATION:
- 80-100%: Excellent - High quality, minimal methodological concerns
- 70-79%:  Good - Good quality with minor limitations  
- 60-69%:  Fair - Moderate quality with some issues
- 0-59%:   Poor - Low quality with significant concerns

USAGE RECOMMENDATIONS:
- Review the PDF report for comprehensive quality analysis
- Use the data file for further statistical analysis or integration
- Consider the quality rating when including in systematic reviews

Generated by AI-Powered Literature Quality Assessment Platform
"""
                readme_path = os.path.join(temp_dir, "README.txt")
                with open(readme_path, 'w', encoding='utf-8') as f:
                    f.write(readme_content)
                zipf.write(readme_path, "README.txt")
            
            # 4. Send ZIP file
            return send_file(
                zip_path,
                mimetype='application/zip',
                as_attachment=True,
                download_name=zip_filename
            )
            
    except Exception as e:
        current_app.logger.error(f"Error generating assessment package for {assessment_id}: {e}")
        flash(f"Error generating assessment package: {str(e)}", 'error')
        return redirect(url_for('.view_assessment_result', assessment_id=assessment_id))

def create_beautiful_pdf_report(pdf_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score):
    """Create a beautiful PDF report with logo and professional formatting"""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing, Rect
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics import renderPDF
    from datetime import datetime
    
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Times New Roman font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#6b46c1'),
        alignment=1,
        fontName='Times-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        textColor=colors.HexColor('#4a5568'),
        fontName='Times-Bold'
    )
    
    # Header with logo only
    try:
        # Add logo
        png_logo_path = '/Users/hongchaokun/Desktop/screen_webapp/app/static/images/Meta_Screener_LOGO.png'
        if os.path.exists(png_logo_path):
            logo = Image(png_logo_path, width=3*inch, height=1.2*inch)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 30))
        else:
            # Fallback to text if no logo image
            story.append(Paragraph("Quality Assessment Report", ParagraphStyle('Brand', parent=styles['Normal'], fontSize=18, textColor=colors.HexColor('#6b46c1'), alignment=1, fontName='Times-Bold')))
            story.append(Spacer(1, 30))
    except Exception as e:
        # Fallback to text if logo loading fails
        story.append(Paragraph("Quality Assessment Report", ParagraphStyle('Brand', parent=styles['Normal'], fontSize=18, textColor=colors.HexColor('#6b46c1'), alignment=1, fontName='Times-Bold')))
        story.append(Spacer(1, 30))
    
    # Title
    story.append(Paragraph("Quality Assessment Report", title_style))
    story.append(Spacer(1, 20))
    
    # Document info table
    doc_info = [
        ['Document Name:', filename],
        ['Assessment ID:', assessment_id],
        ['Document Type:', doc_type],
        ['Assessment Tool:', tool_name],
        ['Report Generated:', datetime.now().strftime('%B %d, %Y at %H:%M:%S')],
    ]
    
    doc_table = Table(doc_info, colWidths=[2.5*inch, 3.5*inch])
    doc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f0ff')),  # Light purple background
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('FONTNAME', (0, 0), (0, -1), 'Times-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('PADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe')),  # Purple border
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b46c1')),  # Purple text for labels
    ]))
    
    story.append(doc_table)
    story.append(Spacer(1, 30))
    
    # Quality score section with visual
    story.append(Paragraph("Assessment Summary", subtitle_style))
    
    # Create quality score visualization
    quality_rating = 'Excellent' if quality_score >= 80 else 'Good' if quality_score >= 70 else 'Fair' if quality_score >= 60 else 'Poor'
    score_color = colors.HexColor('#48bb78') if quality_score >= 80 else colors.HexColor('#4299e1') if quality_score >= 70 else colors.HexColor('#ed8936') if quality_score >= 60 else colors.HexColor('#e53e3e')
    
    summary_data = [
        ['Overall Quality Score:', f"{quality_score}% ({quality_rating})"],
        ['Total Criteria Evaluated:', str(total_criteria)],
        ['Criteria Passed:', str(total_criteria - negative_findings)],
        ['Criteria Failed:', str(negative_findings)],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f0ff')),  # Light purple background
        ('BACKGROUND', (1, 0), (1, 0), score_color),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b46c1')),  # Purple text for labels
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('PADDING', (0, 0), (-1, -1), 15),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe')),  # Purple border
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Skip detailed results as they are available in Excel file
    
    # Add quality score interpretation and recommendations
    story.append(Spacer(1, 30))
    story.append(Paragraph("Quality Score Interpretation & Recommendations", subtitle_style))
    
    # Create compact interpretation table
    interpretation_data = [
        ['Score', 'Rating', 'Interpretation & Recommendation']
    ]
    
    score_ranges = [
        ('80-100%', 'Excellent', 'High quality study with minimal concerns. Include with high confidence.'),
        ('70-79%', 'Good', 'Good quality with minor limitations. Include with moderate confidence.'),
        ('60-69%', 'Fair', 'Moderate quality with some issues. Include with caution, note limitations.'),
        ('0-59%', 'Poor', 'Low quality with significant concerns. Consider exclusion or major caveats.')
    ]
    
    for score_range, rating, interpretation in score_ranges:
        interpretation_data.append([score_range, rating, interpretation])
    
    # Convert text to Paragraph objects for automatic wrapping
    wrapped_interpretation_data = [interpretation_data[0]]  # Keep header as is
    for i, row in enumerate(interpretation_data[1:], 1):
        wrapped_row = [
            row[0],  # Score range - keep as text
            row[1],  # Rating - keep as text
            Paragraph(row[2], ParagraphStyle('CellText', parent=styles['Normal'], fontSize=9, leading=11))  # Wrap long text
        ]
        wrapped_interpretation_data.append(wrapped_row)
    
    interpretation_table = Table(wrapped_interpretation_data, colWidths=[1.2*inch, 1.3*inch, 4*inch])
    interpretation_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b46c1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (1, -1), 9),  # Only apply to non-paragraph cells
        ('PADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        # Enhanced color coding with purple theme
        ('BACKGROUND', (0, 1), (1, 1), colors.HexColor('#d4edda')),  # Excellent - light green
        ('BACKGROUND', (0, 2), (1, 2), colors.HexColor('#cce7ff')),  # Good - light blue  
        ('BACKGROUND', (0, 3), (1, 3), colors.HexColor('#fff0cd')),  # Fair - light yellow
        ('BACKGROUND', (0, 4), (1, 4), colors.HexColor('#ffe6e6')),  # Poor - light red
        # Add subtle purple accent to rating column
        ('TEXTCOLOR', (1, 1), (1, -1), colors.HexColor('#6b46c1')),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(interpretation_table)
    
    # Add specific recommendations based on current assessment quality score
    story.append(Spacer(1, 20))
    
    # Determine current assessment quality rating
    current_rating = 'Poor'
    if quality_score >= 80:
        current_rating = 'Excellent'
        recommendation_text = "This study demonstrates excellent methodological quality. It can be included in your systematic review with high confidence. The findings are likely to be reliable and contribute significantly to your evidence synthesis."
    elif quality_score >= 70:
        current_rating = 'Good'
        recommendation_text = "This study shows good methodological quality with only minor limitations. It should be included in your systematic review with moderate confidence. Consider noting any identified limitations in your quality assessment summary."
    elif quality_score >= 60:
        current_rating = 'Fair'
        recommendation_text = "This study has moderate methodological quality with some concerns. Include it in your review but exercise caution when interpreting results. Clearly document the limitations and consider their impact on your conclusions."
    else:
        current_rating = 'Poor'
        recommendation_text = "This study has significant methodological limitations that may affect the reliability of its findings. Consider excluding it from your primary analysis or include it only in sensitivity analyses with appropriate caveats."
    
    story.append(Paragraph(f"Assessment for This Document ({quality_score}% - {current_rating})", 
                          ParagraphStyle('RecTitle', parent=styles['Heading3'], fontSize=12, textColor=colors.HexColor('#4a5568'), fontName='Times-Bold')))
    story.append(Paragraph(recommendation_text, 
                          ParagraphStyle('RecText', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#2d3748'), leftIndent=20, rightIndent=20, fontName='Times-Roman')))
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph("Generated by AI-Powered Literature Quality Assessment Platform", 
                          ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#718096'), alignment=1)))
    
    doc.build(story)

def create_csv_data_file(csv_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score):
    """Create CSV data file with headers"""
    import pandas as pd
    from datetime import datetime
    
    # Prepare detailed data
    export_data = []
    assessment_details = result_data.get("assessment_details", [])
    
    if isinstance(assessment_details, list) and assessment_details:
        for i, detail in enumerate(assessment_details, 1):
            row_data = {
                'Document_Name': filename,
                'Assessment_ID': assessment_id,
                'Document_Type': doc_type,
                'Assessment_Tool': tool_name,
                'Total_Criteria': total_criteria,
                'Criteria_Passed': total_criteria - negative_findings,
                'Criteria_Failed': negative_findings,
                'Quality_Score_Percent': quality_score,
                'Criterion_Number': i,
                'Criterion_Text': detail.get('criterion_text', 'No criterion text'),
                'AI_Judgment': detail.get('judgment', 'No judgment'),
                'AI_Reasoning': detail.get('reason', 'No reasoning provided'),
                'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            export_data.append(row_data)
    
    df = pd.DataFrame(export_data)
    
    with open(csv_path, 'w', encoding='utf-8-sig') as f:
        # Add comprehensive header information with purple theme
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write("# 🔬 AI-Powered Literature Quality Assessment - Individual Results\n")
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("#\n")
        f.write("# 📊 ASSESSMENT SUMMARY\n")
        f.write(f"# Document Name: {filename}\n")
        f.write(f"# Assessment ID: {assessment_id}\n")
        f.write(f"# Quality Score: {quality_score}%\n")
        f.write(f"# Quality Rating: {'Excellent' if quality_score >= 80 else 'Good' if quality_score >= 70 else 'Fair' if quality_score >= 60 else 'Poor'}\n")
        f.write(f"# Total Criteria: {total_criteria}\n")
        f.write(f"# Criteria Passed: {total_criteria - negative_findings}\n")
        f.write(f"# Criteria Failed: {negative_findings}\n")
        f.write("#\n")
        f.write("# 🎯 QUALITY SCORE INTERPRETATION\n")
        f.write("# 80-100%: Excellent - High quality, minimal concerns\n")
        f.write("# 70-79%:  Good - Good quality, minor limitations\n")
        f.write("# 60-69%:  Fair - Moderate quality, some issues\n")
        f.write("# 0-59%:   Poor - Low quality, significant concerns\n")
        f.write("#\n")
        f.write("# 📋 DETAILED ASSESSMENT DATA\n")
        f.write("# Each row represents one assessment criterion\n")
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write("#\n")
        
        # Write the actual data
        df.to_csv(f, index=False)

def create_excel_data_file(excel_path, result_data, assessment_id, filename, doc_type, tool_name, total_criteria, negative_findings, quality_score):
    """Create Excel data file with formatting"""
    import pandas as pd
    from datetime import datetime
    
    # Prepare detailed data
    export_data = []
    assessment_details = result_data.get("assessment_details", [])
    
    if isinstance(assessment_details, list) and assessment_details:
        for i, detail in enumerate(assessment_details, 1):
            row_data = {
                'Document_Name': filename,
                'Assessment_ID': assessment_id,
                'Document_Type': doc_type,
                'Assessment_Tool': tool_name,
                'Total_Criteria': total_criteria,
                'Criteria_Passed': total_criteria - negative_findings,
                'Criteria_Failed': negative_findings,
                'Quality_Score_Percent': quality_score,
                'Criterion_Number': i,
                'Criterion_Text': detail.get('criterion_text', 'No criterion text'),
                'AI_Judgment': detail.get('judgment', 'No judgment'),
                'AI_Reasoning': detail.get('reason', 'No reasoning provided'),
                'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            export_data.append(row_data)
    
    df = pd.DataFrame(export_data)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Enhanced Summary sheet
        summary_data = {
            'Metric': [
                'Document Name',
                'Assessment ID', 
                'Quality Score',
                'Quality Rating',
                'Total Criteria',
                'Criteria Passed',
                'Criteria Failed',
                'Generated Date'
            ],
            'Value': [
                filename,
                assessment_id,
                f"{quality_score}%",
                'Excellent' if quality_score >= 80 else 'Good' if quality_score >= 70 else 'Fair' if quality_score >= 60 else 'Poor',
                total_criteria,
                total_criteria - negative_findings,
                negative_findings,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name='Summary', startrow=2)
        
        # Detailed assessment data
        df.to_excel(writer, index=False, sheet_name='Detailed Results')
        
        # Quality interpretation guide
        guide_data = {
            'Quality Score Range': ['80-100%', '70-79%', '60-69%', '0-59%'],
            'Rating': ['Excellent', 'Good', 'Fair', 'Poor'],
            'Interpretation': [
                'High quality study with minimal methodological concerns',
                'Good quality study with minor limitations',
                'Moderate quality study with some methodological issues',
                'Low quality study with significant methodological concerns'
            ],
            'Recommendation': [
                'Include in systematic review with high confidence',
                'Include in systematic review with moderate confidence',
                'Include with caution, note limitations in analysis',
                'Consider exclusion or include with major caveats'
            ]
        }
        guide_df = pd.DataFrame(guide_data)
        guide_df.to_excel(writer, index=False, sheet_name='Interpretation Guide')
        
        # Format the Excel file with purple theme
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Format Summary sheet
        summary_ws = writer.sheets['Summary']
        summary_ws['A1'] = 'AI-Powered Literature Quality Assessment - Individual Results'
        summary_ws['A1'].font = Font(name='Times New Roman', size=16, bold=True, color='6b46c1')
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws.merge_cells('A1:B1')
        
        # Header formatting with purple theme and Times New Roman font
        header_fill = PatternFill(start_color='6b46c1', end_color='6b46c1', fill_type='solid')
        header_font = Font(name='Times New Roman', color='FFFFFF', bold=True)
        
        for col in ['A', 'B']:
            cell = summary_ws[f'{col}3']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format Detailed Results sheet
        detail_ws = writer.sheets['Detailed Results']
        
        # Header formatting for detailed results
        for col_num in range(1, len(df.columns) + 1):
            cell = detail_ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format Interpretation Guide sheet
        guide_ws = writer.sheets['Interpretation Guide']
        
        # Header formatting for guide
        for col_num in range(1, len(guide_df.columns) + 1):
            cell = guide_ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Color code the quality ratings in guide (same as individual)
        quality_colors = {
            'Excellent': PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid'),
            'Good': PatternFill(start_color='cce7ff', end_color='cce7ff', fill_type='solid'),
            'Fair': PatternFill(start_color='fff0cd', end_color='fff0cd', fill_type='solid'),
            'Poor': PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')
        }
        
        for row_num in range(2, len(guide_df) + 2):
            rating_cell = guide_ws.cell(row=row_num, column=2)
            rating = rating_cell.value
            if rating in quality_colors:
                for col_num in range(1, len(guide_df.columns) + 1):
                    guide_ws.cell(row=row_num, column=col_num).fill = quality_colors[rating]
        
        # Color code quality scores in detailed results based on Overall_Quality_Score_Percent
        if 'Overall_Quality_Score_Percent' in df.columns:
            score_col_idx = df.columns.get_loc('Overall_Quality_Score_Percent') + 1
            for row_num in range(2, len(df) + 2):
                score_cell = detail_ws.cell(row=row_num, column=score_col_idx)
                try:
                    score_value = float(str(score_cell.value).replace('%', ''))
                    if score_value >= 80:
                        score_cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                    elif score_value >= 70:
                        score_cell.fill = PatternFill(start_color='cce7ff', end_color='cce7ff', fill_type='solid')
                    elif score_value >= 60:
                        score_cell.fill = PatternFill(start_color='fff0cd', end_color='fff0cd', fill_type='solid')
                    else:
                        score_cell.fill = PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')
                except:
                    pass
        
        # Color code the quality ratings in guide
        quality_colors = {
            'Excellent': PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid'),
            'Good': PatternFill(start_color='cce7ff', end_color='cce7ff', fill_type='solid'),
            'Fair': PatternFill(start_color='fff0cd', end_color='fff0cd', fill_type='solid'),
            'Poor': PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')
        }
        
        for row_num in range(2, len(guide_df) + 2):
            rating_cell = guide_ws.cell(row=row_num, column=2)
            rating = rating_cell.value
            if rating in quality_colors:
                for col_num in range(1, len(guide_df.columns) + 1):
                    guide_ws.cell(row=row_num, column=col_num).fill = quality_colors[rating]
        
        # Auto-adjust column widths and set font/alignment
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for col_num in range(1, worksheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Set Times New Roman font and wrap text for all cells
                for row_num in range(1, worksheet.max_row + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value:
                            # Set Times New Roman font for all cells
                            if row_num == 1:  # Header row
                                cell.font = Font(name='Times New Roman', bold=True, color='FFFFFF' if cell.fill.start_color.rgb == 'FF6b46c1' else '000000')
                            else:
                                cell.font = Font(name='Times New Roman')
                            
                            # Enable text wrapping for all cells
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            # Calculate max length for column width
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with special handling for document name columns
                if 'Document' in str(worksheet.cell(row=1, column=col_num).value) or 'Name' in str(worksheet.cell(row=1, column=col_num).value):
                    adjusted_width = min(max_length + 2, 50)  # Limit document name column width
                else:
                    adjusted_width = min(max_length + 2, 80)
                
                if adjusted_width > 8:
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Set row height to accommodate wrapped text
                for row_num in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = None  # Auto-adjust height

# --- NEW: Batch Assessment Package Download --- #
@quality_bp.route('/download_batch_package/<batch_id>/<format>')
def download_batch_package(batch_id, format):
    """Download batch assessment package: data file (Excel/CSV) + beautiful PDF report"""
    import pandas as pd
    import zipfile
    import tempfile
    import os
    from flask import send_file
    from datetime import datetime
    
    batch_info = get_batch_info(batch_id)
    if not batch_info:
        flash("Batch assessment not found.", 'error')
        return redirect(url_for('.upload_document_for_assessment'))
    
    try:
        # Create temporary directory for package files
        with tempfile.TemporaryDirectory() as temp_dir:
            # 1. Generate beautiful batch PDF report
            pdf_filename = f"batch_report_{batch_id[:8]}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            create_batch_pdf_report(pdf_path, batch_info, batch_id)
            
            # 2. Generate batch data file (CSV or Excel)
            if format == 'csv':
                data_filename = f"batch_data_{batch_id[:8]}.csv"
                data_path = os.path.join(temp_dir, data_filename)
                create_batch_csv_data_file(data_path, batch_info, batch_id)
            else:  # xlsx
                data_filename = f"batch_data_{batch_id[:8]}.xlsx"
                data_path = os.path.join(temp_dir, data_filename)
                create_batch_excel_data_file(data_path, batch_info, batch_id)
            
            # 3. Create ZIP package
            zip_filename = f"batch_package_{batch_id[:8]}.zip"
            zip_path = os.path.join(temp_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(pdf_path, pdf_filename)
                zipf.write(data_path, data_filename)
                
                # Add README file
                assessment_ids = batch_info.get("assessment_ids", [])
                completed_count = sum(1 for aid in assessment_ids if get_assessment_result(aid) and get_assessment_result(aid).get("status") == "completed")
                
                readme_content = f"""AI-Powered Literature Quality Assessment - Batch Package
======================================================

This package contains:
1. {pdf_filename} - Professional batch assessment report with quality analysis and recommendations
2. {data_filename} - Comprehensive batch data with all individual assessment details

BATCH SUMMARY:
- Batch ID: {batch_id}
- Total Files: {len(assessment_ids)}
- Completed Assessments: {completed_count}
- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

DATA STRUCTURE:
- Each row in the data file represents one assessment criterion for one document
- Multiple rows per document = multiple criteria evaluated
- Use Assessment_ID to group criteria by document
- Overall_Quality_Score_Percent shows the quality score for each document

QUALITY SCORE INTERPRETATION:
- 80-100%: Excellent - High quality, minimal methodological concerns
- 70-79%:  Good - Good quality with minor limitations
- 60-69%:  Fair - Moderate quality with some issues
- 0-59%:   Poor - Low quality with significant concerns

USAGE RECOMMENDATIONS:
- Review the PDF report for batch overview and quality trends
- Use the data file for detailed analysis of individual assessments
- Filter by Assessment_ID to analyze specific documents
- Consider quality scores when selecting studies for systematic reviews

Generated by AI-Powered Literature Quality Assessment Platform
"""
                readme_path = os.path.join(temp_dir, "README.txt")
                with open(readme_path, 'w', encoding='utf-8') as f:
                    f.write(readme_content)
                zipf.write(readme_path, "README.txt")
            
            # 4. Send ZIP file
            return send_file(
                zip_path,
                mimetype='application/zip',
                as_attachment=True,
                download_name=zip_filename
            )
            
    except Exception as e:
        current_app.logger.error(f"Error generating batch package for {batch_id}: {e}")
        flash(f"Error generating batch package: {str(e)}", 'error')
        return redirect(url_for('.view_batch_results', batch_id=batch_id))

def create_batch_pdf_report(pdf_path, batch_info, batch_id):
    """Create a beautiful batch PDF report with summary and individual results"""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from datetime import datetime
    
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Times New Roman font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#6b46c1'),
        alignment=1,
        fontName='Times-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        textColor=colors.HexColor('#4a5568'),
        fontName='Times-Bold'
    )
    
    # Header with logo only
    try:
        # Add logo
        png_logo_path = '/Users/hongchaokun/Desktop/screen_webapp/app/static/images/Meta_Screener_LOGO.png'
        if os.path.exists(png_logo_path):
            logo = Image(png_logo_path, width=3*inch, height=1.2*inch)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 30))
        else:
            # Fallback to text if no logo image
            story.append(Paragraph("Batch Quality Assessment Report", ParagraphStyle('Brand', parent=styles['Normal'], fontSize=18, textColor=colors.HexColor('#6b46c1'), alignment=1, fontName='Helvetica-Bold')))
            story.append(Spacer(1, 30))
    except Exception as e:
        # Fallback to text if logo loading fails
        story.append(Paragraph("Batch Quality Assessment Report", ParagraphStyle('Brand', parent=styles['Normal'], fontSize=18, textColor=colors.HexColor('#6b46c1'), alignment=1, fontName='Helvetica-Bold')))
        story.append(Spacer(1, 30))
    
    # Title
    story.append(Paragraph("Batch Quality Assessment Report", title_style))
    story.append(Spacer(1, 20))
    
    # Batch info
    assessment_ids = batch_info.get("assessment_ids", [])
    completed_assessments = []
    processing_count = 0
    error_count = 0
    total_quality_scores = []
    
    for assessment_id in assessment_ids:
        result_data = get_assessment_result(assessment_id)
        if result_data:
            status = result_data.get("status", "unknown")
            if status == 'completed':
                completed_assessments.append(result_data)
                total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
                negative_findings = result_data.get("summary_negative_findings", 0)
                if total_criteria > 0:
                    quality_score = round((total_criteria - negative_findings) / total_criteria * 100, 1)
                    total_quality_scores.append(quality_score)
            elif status in ['processing_assessment', 'pending_assessment']:
                processing_count += 1
            else:
                error_count += 1
    
    avg_quality_score = round(sum(total_quality_scores) / len(total_quality_scores), 1) if total_quality_scores else 0
    
    # Batch summary table
    batch_summary = [
        ['Batch ID:', batch_id[:8] + "..."],
        ['Total Files:', str(len(assessment_ids))],
        ['Completed Assessments:', str(len(completed_assessments))],
        ['Processing:', str(processing_count)],
        ['Errors:', str(error_count)],
        ['Average Quality Score:', f"{avg_quality_score}%" if total_quality_scores else "N/A"],
        ['Report Generated:', datetime.now().strftime('%B %d, %Y at %H:%M:%S')],
    ]
    
    batch_table = Table(batch_summary, colWidths=[2.5*inch, 3.5*inch])
    batch_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f0ff')),  # Light purple background
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b46c1')),  # Purple text for labels
        ('FONTNAME', (0, 0), (0, -1), 'Times-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('PADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe')),  # Purple border
    ]))
    
    story.append(batch_table)
    story.append(Spacer(1, 30))
    
    # Individual assessment results
    if completed_assessments:
        story.append(Paragraph("Individual Assessment Results", subtitle_style))
        story.append(Spacer(1, 15))
        
        # Create results table
        table_data = [['Document', 'Quality Score', 'Criteria Passed/Total', 'Rating']]
        
        for result_data in completed_assessments:
            filename = result_data.get('filename', 'Unknown')
            total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
            negative_findings = result_data.get("summary_negative_findings", 0)
            quality_score = round((total_criteria - negative_findings) / total_criteria * 100, 1) if total_criteria > 0 else 0
            quality_rating = 'Excellent' if quality_score >= 80 else 'Good' if quality_score >= 70 else 'Fair' if quality_score >= 60 else 'Poor'
            
            # Truncate filename if too long
            if len(filename) > 40:
                filename = filename[:40] + "..."
            
            table_data.append([
                filename,
                f"{quality_score}%",
                f"{total_criteria - negative_findings}/{total_criteria}",
                quality_rating
            ])
        
        results_table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b46c1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ]))
        
        # Add alternating row colors and quality score colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                results_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa'))
                ]))
        
        story.append(results_table)
    
    # Add quality score interpretation and recommendations (same as individual)
    story.append(Spacer(1, 30))
    story.append(Paragraph("Quality Score Interpretation & Recommendations", subtitle_style))
    
    # Create compact interpretation table
    interpretation_data = [
        ['Score', 'Rating', 'Interpretation & Recommendation']
    ]
    
    score_ranges = [
        ('80-100%', 'Excellent', 'High quality study with minimal concerns. Include with high confidence.'),
        ('70-79%', 'Good', 'Good quality with minor limitations. Include with moderate confidence.'),
        ('60-69%', 'Fair', 'Moderate quality with some issues. Include with caution, note limitations.'),
        ('0-59%', 'Poor', 'Low quality with significant concerns. Consider exclusion or major caveats.')
    ]
    
    for score_range, rating, interpretation in score_ranges:
        interpretation_data.append([score_range, rating, interpretation])
    
    # Convert text to Paragraph objects for automatic wrapping
    wrapped_interpretation_data = [interpretation_data[0]]  # Keep header as is
    for i, row in enumerate(interpretation_data[1:], 1):
        wrapped_row = [
            row[0],  # Score range - keep as text
            row[1],  # Rating - keep as text
            Paragraph(row[2], ParagraphStyle('CellText', parent=styles['Normal'], fontSize=9, leading=11))  # Wrap long text
        ]
        wrapped_interpretation_data.append(wrapped_row)
    
    interpretation_table = Table(wrapped_interpretation_data, colWidths=[1.2*inch, 1.3*inch, 4*inch])
    interpretation_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b46c1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (1, -1), 9),  # Only apply to non-paragraph cells
        ('PADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        # Enhanced color coding with purple theme
        ('BACKGROUND', (0, 1), (1, 1), colors.HexColor('#d4edda')),  # Excellent - light green
        ('BACKGROUND', (0, 2), (1, 2), colors.HexColor('#cce7ff')),  # Good - light blue  
        ('BACKGROUND', (0, 3), (1, 3), colors.HexColor('#fff0cd')),  # Fair - light yellow
        ('BACKGROUND', (0, 4), (1, 4), colors.HexColor('#ffe6e6')),  # Poor - light red
        # Add subtle purple accent to rating column
        ('TEXTCOLOR', (1, 1), (1, -1), colors.HexColor('#6b46c1')),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(interpretation_table)
    
    # Add batch quality interpretation
    story.append(Spacer(1, 30))
    story.append(Paragraph("Batch Quality Assessment Summary", subtitle_style))
    
    # Overall batch recommendation
    if avg_quality_score >= 80:
        batch_recommendation = "This batch demonstrates excellent overall quality. Most studies can be included in systematic reviews with high confidence."
    elif avg_quality_score >= 70:
        batch_recommendation = "This batch shows good overall quality. Most studies are suitable for inclusion with moderate confidence."
    elif avg_quality_score >= 60:
        batch_recommendation = "This batch has moderate quality. Review individual assessments carefully and consider limitations in your analysis."
    else:
        batch_recommendation = "This batch has significant quality concerns. Careful individual review is recommended, and consider excluding low-quality studies."
    
    story.append(Paragraph(f"Overall Batch Quality: {avg_quality_score}%", 
                          ParagraphStyle('BatchScore', parent=styles['Heading3'], fontSize=12, textColor=colors.HexColor('#4a5568'), fontName='Times-Bold')))
    story.append(Paragraph(batch_recommendation, 
                          ParagraphStyle('BatchRec', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#2d3748'), leftIndent=20, rightIndent=20, fontName='Times-Roman')))
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph("Generated by AI-Powered Literature Quality Assessment Platform", 
                          ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#718096'), alignment=1)))
    
    doc.build(story)

def create_batch_csv_data_file(csv_path, batch_info, batch_id):
    """Create comprehensive batch CSV data file with all assessment details"""
    import pandas as pd
    from datetime import datetime
    
    assessment_ids = batch_info.get("assessment_ids", [])
    export_data = []
    
    for idx, assessment_id in enumerate(assessment_ids):
        result_data = get_assessment_result(assessment_id)
        if result_data:
            filename = result_data.get("filename", f"Unknown_{idx}")
            doc_type = result_data.get("document_type", "Unknown")
            status = result_data.get("status", "unknown")
            tool_name = QUALITY_ASSESSMENT_TOOLS.get(doc_type, {}).get('tool_name', 'Standard Assessment')
            
            if status == 'completed':
                total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
                negative_findings = result_data.get("summary_negative_findings", 0)
                quality_score = round((total_criteria - negative_findings) / total_criteria * 100, 1) if total_criteria > 0 else 0
                
                # Get detailed assessment results
                assessment_details = result_data.get("assessment_details", [])
                
                if isinstance(assessment_details, list) and assessment_details:
                    # Create one row per criterion
                    for i, detail in enumerate(assessment_details, 1):
                        row_data = {
                            'Batch_ID': batch_id,
                            'Assessment_ID': assessment_id,
                            'Document_Name': filename,
                            'Document_Type': doc_type,
                            'Assessment_Tool': tool_name,
                            'Overall_Quality_Score_Percent': quality_score,
                            'Total_Criteria': total_criteria,
                            'Criteria_Passed': total_criteria - negative_findings,
                            'Criteria_Failed': negative_findings,
                            'Criterion_Number': i,
                            'Criterion_ID': detail.get('criterion_id', f'criterion_{i}'),
                            'Criterion_Text': detail.get('criterion_text', 'No criterion text'),
                            'AI_Judgment': detail.get('judgment', 'No judgment'),
                            'AI_Reasoning': detail.get('reason', 'No reasoning provided'),
                            'Status': status,
                            'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        export_data.append(row_data)
                else:
                    # If no detailed results, create summary row
                    row_data = {
                        'Batch_ID': batch_id,
                        'Assessment_ID': assessment_id,
                        'Document_Name': filename,
                        'Document_Type': doc_type,
                        'Assessment_Tool': tool_name,
                        'Overall_Quality_Score_Percent': quality_score,
                        'Total_Criteria': total_criteria,
                        'Criteria_Passed': total_criteria - negative_findings,
                        'Criteria_Failed': negative_findings,
                        'Criterion_Number': 'N/A',
                        'Criterion_ID': 'N/A',
                        'Criterion_Text': 'No detailed assessment available',
                        'AI_Judgment': 'N/A',
                        'AI_Reasoning': 'No detailed reasoning available',
                        'Status': status,
                        'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    export_data.append(row_data)
            else:
                # For non-completed assessments
                row_data = {
                    'Batch_ID': batch_id,
                    'Assessment_ID': assessment_id,
                    'Document_Name': filename,
                    'Document_Type': doc_type,
                    'Assessment_Tool': tool_name,
                    'Overall_Quality_Score_Percent': 'N/A',
                    'Total_Criteria': 'N/A',
                    'Criteria_Passed': 'N/A',
                    'Criteria_Failed': 'N/A',
                    'Criterion_Number': 'N/A',
                    'Criterion_ID': 'N/A',
                    'Criterion_Text': 'Assessment not completed',
                    'AI_Judgment': 'N/A',
                    'AI_Reasoning': result_data.get('message', 'Assessment in progress or failed'),
                    'Status': status,
                    'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                export_data.append(row_data)
    
    df = pd.DataFrame(export_data)
    
    with open(csv_path, 'w', encoding='utf-8-sig') as f:
        # Add comprehensive header information with beautiful formatting
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write("# 🔬 AI-Powered Literature Quality Assessment - Batch Results\n")
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"# Batch ID: {batch_id}\n")
        f.write(f"# Total Assessments: {len(assessment_ids)}\n")
        
        # Calculate batch statistics
        completed_count = sum(1 for row in export_data if row.get('Status') == 'completed')
        if completed_count > 0:
            avg_score = sum(row.get('Overall_Quality_Score_Percent', 0) for row in export_data if isinstance(row.get('Overall_Quality_Score_Percent'), (int, float))) / completed_count
            f.write(f"# Completed Assessments: {completed_count}\n")
            f.write(f"# Average Quality Score: {avg_score:.1f}%\n")
        
        f.write("#\n")
        f.write("# 📊 DATA STRUCTURE\n")
        f.write("# Each row represents one assessment criterion for one document\n")
        f.write("# Multiple rows per document = multiple criteria evaluated\n")
        f.write("# Use Assessment_ID to group criteria by document\n")
        f.write("# Overall_Quality_Score_Percent shows the quality score for each document\n")
        f.write("#\n")
        f.write("# 🎯 QUALITY SCORE INTERPRETATION\n")
        f.write("# 80-100%: Excellent - High quality, minimal methodological concerns\n")
        f.write("# 70-79%:  Good - Good quality with minor limitations\n")
        f.write("# 60-69%:  Fair - Moderate quality with some issues\n")
        f.write("# 0-59%:   Poor - Low quality with significant concerns\n")
        f.write("#\n")
        f.write("# 💡 USAGE RECOMMENDATIONS\n")
        f.write("# - Filter by Assessment_ID to analyze specific documents\n")
        f.write("# - Sort by Overall_Quality_Score_Percent to prioritize high-quality studies\n")
        f.write("# - Use AI_Judgment and AI_Reasoning for detailed quality analysis\n")
        f.write("# - Consider quality scores when selecting studies for systematic reviews\n")
        f.write("#\n")
        f.write("# 📋 DETAILED BATCH ASSESSMENT DATA\n")
        f.write("# ═══════════════════════════════════════════════════════════════\n")
        f.write("#\n")
        
        # Write the actual data
        df.to_csv(f, index=False)

def create_batch_excel_data_file(excel_path, batch_info, batch_id):
    """Create comprehensive batch Excel data file with all assessment details and formatting"""
    import pandas as pd
    from datetime import datetime
    
    assessment_ids = batch_info.get("assessment_ids", [])
    export_data = []
    summary_stats = {'completed': 0, 'total_scores': []}
    
    for idx, assessment_id in enumerate(assessment_ids):
        result_data = get_assessment_result(assessment_id)
        if result_data:
            filename = result_data.get("filename", f"Unknown_{idx}")
            doc_type = result_data.get("document_type", "Unknown")
            status = result_data.get("status", "unknown")
            tool_name = QUALITY_ASSESSMENT_TOOLS.get(doc_type, {}).get('tool_name', 'Standard Assessment')
            
            if status == 'completed':
                total_criteria = result_data.get("summary_total_criteria_evaluated", 0)
                negative_findings = result_data.get("summary_negative_findings", 0)
                quality_score = round((total_criteria - negative_findings) / total_criteria * 100, 1) if total_criteria > 0 else 0
                
                summary_stats['completed'] += 1
                summary_stats['total_scores'].append(quality_score)
                
                # Get detailed assessment results
                assessment_details = result_data.get("assessment_details", [])
                
                if isinstance(assessment_details, list) and assessment_details:
                    # Create one row per criterion
                    for i, detail in enumerate(assessment_details, 1):
                        row_data = {
                            'Batch_ID': batch_id,
                            'Assessment_ID': assessment_id,
                            'Document_Name': filename,
                            'Document_Type': doc_type,
                            'Assessment_Tool': tool_name,
                            'Overall_Quality_Score_Percent': quality_score,
                            'Total_Criteria': total_criteria,
                            'Criteria_Passed': total_criteria - negative_findings,
                            'Criteria_Failed': negative_findings,
                            'Criterion_Number': i,
                            'Criterion_ID': detail.get('criterion_id', f'criterion_{i}'),
                            'Criterion_Text': detail.get('criterion_text', 'No criterion text'),
                            'AI_Judgment': detail.get('judgment', 'No judgment'),
                            'AI_Reasoning': detail.get('reason', 'No reasoning provided'),
                            'Status': status,
                            'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        export_data.append(row_data)
                else:
                    # If no detailed results, create summary row
                    row_data = {
                        'Batch_ID': batch_id,
                        'Assessment_ID': assessment_id,
                        'Document_Name': filename,
                        'Document_Type': doc_type,
                        'Assessment_Tool': tool_name,
                        'Overall_Quality_Score_Percent': quality_score,
                        'Total_Criteria': total_criteria,
                        'Criteria_Passed': total_criteria - negative_findings,
                        'Criteria_Failed': negative_findings,
                        'Criterion_Number': 'N/A',
                        'Criterion_ID': 'N/A',
                        'Criterion_Text': 'No detailed assessment available',
                        'AI_Judgment': 'N/A',
                        'AI_Reasoning': 'No detailed reasoning available',
                        'Status': status,
                        'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    export_data.append(row_data)
            else:
                # For non-completed assessments
                row_data = {
                    'Batch_ID': batch_id,
                    'Assessment_ID': assessment_id,
                    'Document_Name': filename,
                    'Document_Type': doc_type,
                    'Assessment_Tool': tool_name,
                    'Overall_Quality_Score_Percent': 'N/A',
                    'Total_Criteria': 'N/A',
                    'Criteria_Passed': 'N/A',
                    'Criteria_Failed': 'N/A',
                    'Criterion_Number': 'N/A',
                    'Criterion_ID': 'N/A',
                    'Criterion_Text': 'Assessment not completed',
                    'AI_Judgment': 'N/A',
                    'AI_Reasoning': result_data.get('message', 'Assessment in progress or failed'),
                    'Status': status,
                    'Generated_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                export_data.append(row_data)
    
    df = pd.DataFrame(export_data)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Enhanced Summary sheet
        avg_score = sum(summary_stats['total_scores']) / max(len(summary_stats['total_scores']), 1) if summary_stats['total_scores'] else 0
        
        summary_data = {
            'Metric': [
                'Batch ID',
                'Total Files',
                'Completed Assessments',
                'Processing/Failed',
                'Average Quality Score',
                'Quality Rating',
                'Generated Date'
            ],
            'Value': [
                batch_id,
                len(assessment_ids),
                summary_stats['completed'],
                len(assessment_ids) - summary_stats['completed'],
                f"{avg_score:.1f}%",
                'Excellent' if avg_score >= 80 else 'Good' if avg_score >= 70 else 'Fair' if avg_score >= 60 else 'Poor',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name='Batch Summary', startrow=2)
        
        # Detailed assessment data
        df.to_excel(writer, index=False, sheet_name='Detailed Results')
        
        # Quality interpretation guide
        guide_data = {
            'Quality Score Range': ['80-100%', '70-79%', '60-69%', '0-59%'],
            'Rating': ['Excellent', 'Good', 'Fair', 'Poor'],
            'Interpretation': [
                'High quality studies with minimal methodological concerns',
                'Good quality studies with minor limitations',
                'Moderate quality studies with some methodological issues',
                'Low quality studies with significant methodological concerns'
            ],
            'Recommendation': [
                'Include in systematic review with high confidence',
                'Include in systematic review with moderate confidence',
                'Include with caution, note limitations in analysis',
                'Consider exclusion or include with major caveats'
            ]
        }
        guide_df = pd.DataFrame(guide_data)
        guide_df.to_excel(writer, index=False, sheet_name='Interpretation Guide')
        
        # Format the Excel file with Times New Roman font
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Format Summary sheet
        summary_ws = writer.sheets['Batch Summary']
        summary_ws['A1'] = 'AI-Powered Literature Quality Assessment - Batch Results'
        summary_ws['A1'].font = Font(name='Times New Roman', size=16, bold=True, color='6b46c1')
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws.merge_cells('A1:B1')
        
        # Header formatting with Times New Roman
        header_fill = PatternFill(start_color='6b46c1', end_color='6b46c1', fill_type='solid')
        header_font = Font(name='Times New Roman', color='FFFFFF', bold=True)
        
        for col in ['A', 'B']:
            cell = summary_ws[f'{col}3']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format Detailed Results sheet
        detail_ws = writer.sheets['Detailed Results']
        
        # Header formatting for detailed results
        for col_num in range(1, len(df.columns) + 1):
            cell = detail_ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format Interpretation Guide sheet
        guide_ws = writer.sheets['Interpretation Guide']
        
        # Header formatting for guide
        for col_num in range(1, len(guide_df.columns) + 1):
            cell = guide_ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Color code the quality ratings in guide (same as individual)
        quality_colors = {
            'Excellent': PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid'),
            'Good': PatternFill(start_color='cce7ff', end_color='cce7ff', fill_type='solid'),
            'Fair': PatternFill(start_color='fff0cd', end_color='fff0cd', fill_type='solid'),
            'Poor': PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')
        }
        
        for row_num in range(2, len(guide_df) + 2):
            rating_cell = guide_ws.cell(row=row_num, column=2)
            rating = rating_cell.value
            if rating in quality_colors:
                for col_num in range(1, len(guide_df.columns) + 1):
                    guide_ws.cell(row=row_num, column=col_num).fill = quality_colors[rating]
        
        # Auto-adjust column widths and set font/alignment
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for col_num in range(1, worksheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Set Times New Roman font and wrap text for all cells
                for row_num in range(1, worksheet.max_row + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value:
                            # Set Times New Roman font for all cells
                            if row_num == 1:  # Header row
                                cell.font = Font(name='Times New Roman', bold=True, color='FFFFFF' if cell.fill.start_color.rgb == 'FF6b46c1' else '000000')
                            else:
                                cell.font = Font(name='Times New Roman')
                            
                            # Enable text wrapping for all cells
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            # Calculate max length for column width
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with special handling for document name columns
                if 'Document' in str(worksheet.cell(row=1, column=col_num).value) or 'Name' in str(worksheet.cell(row=1, column=col_num).value):
                    adjusted_width = min(max_length + 2, 50)  # Limit document name column width
                else:
                    adjusted_width = min(max_length + 2, 80)
                
                if adjusted_width > 8:
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Set row height to accommodate wrapped text
                for row_num in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = None  # Auto-adjust height

# End of routes file