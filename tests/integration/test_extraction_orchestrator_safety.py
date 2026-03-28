"""Safety-net integration tests for the extraction orchestrator and exporter.

These tests verify that extract_pdf() and export_to_excel() work end-to-end
with a mock LLM backend.  They serve as a baseline regression guard before
any refactoring of the extraction engine.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from metascreener.core.enums import FieldRole, SheetCardinality, SheetRole
from metascreener.core.models_extraction import (
    ExtractionSchema,
    ExtractionSessionResult,
    FieldSchema,
    SheetSchema,
)
from metascreener.module2_extraction.engine.orchestrator import extract_pdf
from metascreener.module2_extraction.exporter import export_to_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_simple_schema() -> ExtractionSchema:
    """Build a minimal ExtractionSchema with one DATA sheet and two fields."""
    fields = [
        FieldSchema(
            column="A",
            name="Study ID",
            description="Unique study identifier",
            field_type="text",
            role=FieldRole.EXTRACT,
            required=True,
        ),
        FieldSchema(
            column="B",
            name="Sample Size",
            description="Total number of participants",
            field_type="integer",
            role=FieldRole.EXTRACT,
            required=False,
        ),
    ]
    sheet = SheetSchema(
        sheet_name="Studies",
        role=SheetRole.DATA,
        cardinality=SheetCardinality.ONE_PER_STUDY,
        fields=fields,
        extraction_order=1,
    )
    return ExtractionSchema(
        schema_id="test-schema-001",
        schema_version="1.0.0",
        sheets=[sheet],
    )


class MockLLMBackend:
    """Minimal async LLM backend that returns canned extraction JSON.

    Implements the single method the orchestrator calls:
        await backend.complete(prompt, seed=42) -> str
    """

    model_id = "mock-llm"

    async def complete(self, prompt: str, *, seed: int = 42) -> str:  # noqa: ARG002
        """Return a fixed JSON payload regardless of the prompt."""
        payload = {
            "extracted_fields": {
                "Study ID": "Smith 2023",
                "Sample Size": 120,
            },
            "evidence": {
                "Study ID": "First author listed as Smith, published 2023.",
                "Sample Size": "120 participants enrolled according to Table 1.",
            },
        }
        return json.dumps(payload)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_orchestrator_produces_results() -> None:
    """extract_pdf() returns a non-None ExtractionSessionResult with sheets."""
    schema = _make_simple_schema()
    backend = MockLLMBackend()

    result = await extract_pdf(
        schema=schema,
        text="Smith 2023 enrolled 120 patients.",
        pdf_id="pdf-001",
        pdf_filename="smith2023.pdf",
        backend_a=backend,
        backend_b=backend,
    )

    assert result is not None, "extract_pdf() returned None"
    assert isinstance(result, ExtractionSessionResult)
    assert result.pdf_id == "pdf-001"
    assert result.pdf_filename == "smith2023.pdf"
    assert len(result.sheets) > 0, "No sheets in result"
    assert "Studies" in result.sheets, (
        f"Expected 'Studies' sheet; got {list(result.sheets.keys())}"
    )

    studies = result.sheets["Studies"]
    assert len(studies.rows) > 0, "Studies sheet has no rows"


@pytest.mark.asyncio
async def test_exporter_produces_excel(tmp_path: Path) -> None:
    """export_to_excel() writes a valid Excel file with expected sheets."""
    schema = _make_simple_schema()
    backend = MockLLMBackend()

    result = await extract_pdf(
        schema=schema,
        text="Smith 2023 enrolled 120 patients.",
        pdf_id="pdf-001",
        pdf_filename="smith2023.pdf",
        backend_a=backend,
        backend_b=backend,
    )

    output_path = tmp_path / "output.xlsx"
    returned_path = export_to_excel(
        schema=schema,
        results=[result],
        output_path=output_path,
    )

    # File must exist at the returned path
    assert returned_path.exists(), f"Excel file not found at {returned_path}"
    assert returned_path == output_path

    # Verify sheet names using openpyxl
    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    sheet_names = wb.sheetnames

    assert "Studies" in sheet_names, (
        f"Expected 'Studies' sheet in Excel; found {sheet_names}"
    )
    assert "_MetaScreener_Log" in sheet_names, (
        f"Expected '_MetaScreener_Log' sheet in Excel; found {sheet_names}"
    )

    # Verify the Studies sheet has a header row with the expected columns
    ws = wb["Studies"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 3)]
    assert "Study ID" in headers, f"Header 'Study ID' missing; got {headers}"
    assert "Sample Size" in headers, f"Header 'Sample Size' missing; got {headers}"

    wb.close()
