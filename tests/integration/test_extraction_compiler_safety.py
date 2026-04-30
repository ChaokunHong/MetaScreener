"""Safety-net integration tests for the extraction compiler pipeline.

These tests verify that the existing compile_template() pipeline works correctly
against real Excel files built with openpyxl.  They MUST pass against the current
codebase and serve as a baseline before any refactoring.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from metascreener.module2_extraction.compiler.compiler import compile_template


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_single_sheet_workbook(path: Path) -> None:
    """Create a minimal one-sheet Excel template and save it to *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Studies"
    ws.append(["Study ID", "Sample Size", "Intervention", "Outcome"])
    wb.save(path)
    wb.close()


def _make_two_sheet_workbook(path: Path) -> None:
    """Create a two-sheet Excel template (Studies + Outcomes) and save it."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Studies"
    ws1.append(["Study ID", "Author", "Year"])

    ws2 = wb.create_sheet("Outcomes")
    ws2.append(["Study ID", "Outcome", "Effect Size"])

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_compiler_produces_valid_schema(tmp_path: Path) -> None:
    """compile_template returns a well-formed ExtractionSchema for a minimal template."""
    template = tmp_path / "minimal.xlsx"
    _make_single_sheet_workbook(template)

    schema = await compile_template(template, llm_backend=None)

    # Schema must not be None and must carry basic identity fields.
    assert schema is not None
    assert schema.schema_id is not None
    assert schema.schema_version is not None

    # At least one sheet must be present.
    assert len(schema.sheets) > 0, "Schema has no sheets"

    # The Studies sheet must appear.
    sheet_names = [s.sheet_name for s in schema.sheets]
    assert "Studies" in sheet_names, f"Expected 'Studies' in {sheet_names}"

    # All expected column names must appear as fields.
    studies_sheet = next(s for s in schema.sheets if s.sheet_name == "Studies")
    field_names = [f.name for f in studies_sheet.fields]
    for expected in ("Study ID", "Sample Size", "Intervention", "Outcome"):
        assert expected in field_names, (
            f"Expected field '{expected}' missing from Studies sheet; got {field_names}"
        )


@pytest.mark.asyncio
async def test_compiler_handles_multi_sheet(tmp_path: Path) -> None:
    """compile_template handles a two-sheet template and returns both sheets."""
    template = tmp_path / "multi_sheet.xlsx"
    _make_two_sheet_workbook(template)

    schema = await compile_template(template, llm_backend=None)

    assert schema is not None
    assert len(schema.sheets) >= 2, (
        f"Expected at least 2 sheets, got {len(schema.sheets)}"
    )

    sheet_names = [s.sheet_name for s in schema.sheets]
    assert "Studies" in sheet_names, f"Expected 'Studies' in {sheet_names}"
    assert "Outcomes" in sheet_names, f"Expected 'Outcomes' in {sheet_names}"

    # Verify that both sheets have the correct fields.
    studies_sheet = next(s for s in schema.sheets if s.sheet_name == "Studies")
    outcomes_sheet = next(s for s in schema.sheets if s.sheet_name == "Outcomes")

    studies_fields = [f.name for f in studies_sheet.fields]
    for expected in ("Study ID", "Author", "Year"):
        assert expected in studies_fields, (
            f"Expected field '{expected}' in Studies; got {studies_fields}"
        )

    outcomes_fields = [f.name for f in outcomes_sheet.fields]
    for expected in ("Study ID", "Outcome", "Effect Size"):
        assert expected in outcomes_fields, (
            f"Expected field '{expected}' in Outcomes; got {outcomes_fields}"
        )
