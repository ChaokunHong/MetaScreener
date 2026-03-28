"""End-to-end integration tests for the new extraction pipeline."""
from __future__ import annotations

import json
import pytest
from pathlib import Path

from metascreener.doc_engine.parser import DocumentParser
from metascreener.core.enums import FieldRole
from metascreener.core.models_extraction import ExtractionSchema, FieldSchema, SheetSchema
from metascreener.module2_extraction.engine.new_orchestrator import NewOrchestrator
from metascreener.module2_extraction.models import ExtractionStrategy


class MockOCRRouter:
    """Returns pre-built markdown simulating a parsed PDF."""

    async def convert_pdf(self, pdf_path: Path):
        from metascreener.module0_retrieval.models import OCRResult

        return OCRResult(
            record_id=pdf_path.stem,
            markdown=(
                "# A Randomized Trial of Drug X vs Placebo\n\n"
                "DOI: 10.1016/test-001\n\n"
                "# Abstract\n"
                "This double-blind RCT enrolled 120 patients.\n\n"
                "# Methods\n"
                "We randomized patients to Drug X (n=60) or placebo (n=60). "
                "The primary outcome was mortality at 30 days.\n\n"
                "**Table 1. Baseline characteristics**\n\n"
                "| Variable | Drug X (n=60) | Placebo (n=60) |\n"
                "|---|---|---|\n"
                "| Age, years | 55.2 | 54.8 |\n"
                "| Male, n (%) | 35 (58.3) | 33 (55.0) |\n"
                "| BMI | 27.1 | 26.8 |\n\n"
                "# Results\n"
                "The primary outcome occurred in 5 of 60 patients (8.3%) in the "
                "Drug X group and 12 of 60 patients (20.0%) in the placebo group. "
                "The odds ratio was 0.36 (95% CI 0.12-1.10, p=0.07).\n\n"
                "**Table 2. Primary outcome**\n\n"
                "| Outcome | Drug X | Placebo | OR (95% CI) | p |\n"
                "|---|---|---|---|---|\n"
                "| Mortality | 5/60 (8.3%) | 12/60 (20.0%) | 0.36 (0.12-1.10) | 0.07 |\n\n"
                "# Discussion\n"
                "Our findings suggest Drug X may reduce mortality, though the difference "
                "was not statistically significant (p=0.07).\n\n"
                "# References\n"
                "1. Smith J. Prior study. Lancet. 2019.\n"
            ),
            total_pages=8,
            backend_usage={"pymupdf": 8},
            conversion_time_s=1.0,
        )


class MockLLMBackend:
    """Returns canned extraction responses."""

    def __init__(self, model_id: str = "mock") -> None:
        self.model_id = model_id

    async def complete(self, prompt: str, *, seed: int = 42) -> str:
        return json.dumps(
            {
                "fields": {
                    "Study Conclusion": {
                        "value": "Drug X may reduce mortality but not statistically significant",
                        "evidence": (
                            "Our findings suggest Drug X may reduce mortality, "
                            "though the difference was not statistically significant"
                        ),
                    },
                    "Study Design": {
                        "value": "Double-blind RCT",
                        "evidence": "This double-blind RCT enrolled 120 patients",
                    },
                }
            }
        )


def _make_schema(field_names: list[str]) -> ExtractionSchema:
    fields = [
        FieldSchema(
            column=chr(65 + i),
            name=name,
            description=name,
            field_type="text",
            role=FieldRole.EXTRACT,
            required=False,
            dropdown_options=None,
            validation=None,
            mapping_source=None,
        )
        for i, name in enumerate(field_names)
    ]
    sheet = SheetSchema(
        sheet_name="Studies",
        role="data",
        cardinality="one_per_study",
        fields=fields,
        extraction_order=1,
    )
    return ExtractionSchema(
        schema_id="e2e-test",
        schema_version="1.0",
        sheets=[sheet],
        relationships=[],
        mappings={},
        domain_plugin=None,
    )


@pytest.mark.asyncio
async def test_full_pipeline_simple_rct(tmp_path: Path) -> None:
    """End-to-end: PDF parse → field routing → extraction → results."""
    # Step 1: Parse document
    parser = DocumentParser(ocr_router=MockOCRRouter())
    pdf_path = tmp_path / "smith2020.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")
    doc = await parser.parse(pdf_path)

    # Verify DocEngine output
    assert len(doc.sections) >= 4
    assert len(doc.tables) >= 1
    assert doc.metadata.doi == "10.1016/test-001"

    # Step 2: Extract with mixed strategies
    # "Variable" is a column header in Table 1 → DIRECT_TABLE route
    # "Study Conclusion" has no table header match → LLM_TEXT route
    schema = _make_schema(["Variable", "Study Conclusion", "Study Design"])
    orch = NewOrchestrator()
    backend = MockLLMBackend()

    result = await orch.extract(schema, doc, backend, backend)

    # Verify results
    assert result.doc_id is not None
    assert len(result.fields) >= 1

    # "Variable" matches the first column header of Table 1 → DIRECT_TABLE
    if "Variable" in result.fields:
        var_field = result.fields["Variable"]
        assert var_field.strategy == ExtractionStrategy.DIRECT_TABLE
        assert var_field.value is not None

    # Study Conclusion should be LLM_TEXT
    if "Study Conclusion" in result.fields:
        conclusion = result.fields["Study Conclusion"]
        assert conclusion.strategy == ExtractionStrategy.LLM_TEXT
        assert conclusion.value is not None


@pytest.mark.asyncio
async def test_docengine_table_extraction(tmp_path: Path) -> None:
    """Verify DocEngine correctly extracts table structure."""
    parser = DocumentParser(ocr_router=MockOCRRouter())
    pdf_path = tmp_path / "test.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")
    doc = await parser.parse(pdf_path)

    # Should have at least Table 1 (Baseline)
    table1 = doc.get_table("table_1")
    assert table1 is not None
    assert "Baseline" in table1.caption or "characteristics" in table1.caption.lower()

    # Table structure
    assert table1.header_rows == 1
    assert len(table1.cells) >= 4  # header + 3 data rows

    # Header cells should be marked
    for cell in table1.cells[0]:
        assert cell.is_header is True


@pytest.mark.asyncio
async def test_direct_table_vs_llm_routing(tmp_path: Path) -> None:
    """Fields matching tables route to DIRECT_TABLE, others to LLM_TEXT."""
    parser = DocumentParser(ocr_router=MockOCRRouter())
    pdf_path = tmp_path / "test.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")
    doc = await parser.parse(pdf_path)

    from metascreener.module2_extraction.engine.field_router import FieldRouter

    # "Variable" is a column header in Table 1 → routed to DIRECT_TABLE
    # "Study Conclusion" has no table header match → routed to LLM_TEXT
    table_field = FieldSchema(
        column="A",
        name="Variable",
        description="",
        field_type="text",
        role=FieldRole.EXTRACT,
        required=False,
    )
    text_field = FieldSchema(
        column="B",
        name="Study Conclusion",
        description="",
        field_type="text",
        role=FieldRole.EXTRACT,
        required=False,
    )

    router = FieldRouter()
    plans = router.route([table_field, text_field], doc)

    strategies = {p.field_name: p.strategy for p in plans}
    assert strategies.get("Variable") == ExtractionStrategy.DIRECT_TABLE
    assert strategies.get("Study Conclusion") == ExtractionStrategy.LLM_TEXT


@pytest.mark.asyncio
async def test_to_markdown_backward_compat(tmp_path: Path) -> None:
    """StructuredDocument.to_markdown() works for module1 compatibility."""
    parser = DocumentParser(ocr_router=MockOCRRouter())
    pdf_path = tmp_path / "test.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")
    doc = await parser.parse(pdf_path)

    md = doc.to_markdown(strip_references=True)
    assert "Methods" in md
    assert "Results" in md
    assert "References" not in md
    assert len(md) > 100


@pytest.mark.asyncio
async def test_export_roundtrip(tmp_path: Path) -> None:
    """Extract results can be exported to Excel and read back."""
    from metascreener.module2_extraction.export.excel import export_extraction_results

    # Simulate results from repository format
    results = [
        {
            "pdf_id": "pdf1",
            "field_name": "Age",
            "value": "55.2",
            "confidence": "verified",
            "evidence_json": "{}",
            "strategy": "direct_table",
        },
        {
            "pdf_id": "pdf1",
            "field_name": "Study Design",
            "value": "RCT",
            "confidence": "high",
            "evidence_json": "{}",
            "strategy": "llm_text",
        },
    ]

    output = tmp_path / "export.xlsx"
    export_extraction_results(results, ["Age", "Study Design"], output)

    import openpyxl

    wb = openpyxl.load_workbook(output)
    assert "Extraction Results" in wb.sheetnames
    ws = wb["Extraction Results"]
    assert ws.cell(1, 1).value == "Age"
    assert ws.cell(2, 1).value == "55.2"
