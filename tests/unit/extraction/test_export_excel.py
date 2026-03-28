"""Tests for enhanced Excel exporter with confidence color coding."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from metascreener.module2_extraction.export.excel import (
    CONFIDENCE_COLORS,
    export_extraction_results,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_results() -> list[dict]:
    """Build minimal cell-dict list for two PDFs."""
    return [
        {
            "pdf_id": "pdf_001",
            "field_name": "Sample Size",
            "value": "120",
            "confidence": "high",
            "evidence_json": '{"page": 3, "text": "120 patients"}',
        },
        {
            "pdf_id": "pdf_001",
            "field_name": "Intervention",
            "value": "Drug A",
            "confidence": "verified",
            "evidence_json": '{"page": 1, "text": "Drug A was given"}',
        },
        {
            "pdf_id": "pdf_002",
            "field_name": "Sample Size",
            "value": "45",
            "confidence": "medium",
            "evidence_json": '{"page": 2, "text": "45 subjects"}',
        },
        {
            "pdf_id": "pdf_002",
            "field_name": "Intervention",
            "value": "Placebo",
            "confidence": "low",
            "evidence_json": "",
        },
    ]


FIELD_NAMES = ["Sample Size", "Intervention"]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_export_basic(tmp_path: Path) -> None:
    """Export 2 PDFs; verify file exists and headers/data are correct."""
    output = tmp_path / "results.xlsx"
    returned = export_extraction_results(
        results=_make_results(),
        field_names=FIELD_NAMES,
        output_path=output,
        include_evidence=False,
        include_confidence=False,
    )

    assert returned == output
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    ws = wb["Extraction Results"]

    # Headers
    assert ws.cell(1, 1).value == "Sample Size"
    assert ws.cell(1, 2).value == "Intervention"

    # Two data rows
    values = {
        (ws.cell(r, 1).value, ws.cell(r, 2).value)
        for r in (2, 3)
    }
    assert ("120", "Drug A") in values
    assert ("45", "Placebo") in values


def test_export_confidence_colors(tmp_path: Path) -> None:
    """Cells should be color-coded by confidence level."""
    output = tmp_path / "colored.xlsx"
    export_extraction_results(
        results=_make_results(),
        field_names=FIELD_NAMES,
        output_path=output,
        include_evidence=False,
        include_confidence=True,
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Extraction Results"]

    # Build a value→fill map from data rows
    fill_by_value: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value and cell.fill and cell.fill.fgColor:
                color = cell.fill.fgColor.rgb
                # openpyxl returns full ARGB; strip alpha prefix
                rgb = color[-6:] if len(color) == 8 else color
                fill_by_value[str(cell.value)] = rgb.lower()

    assert fill_by_value.get("120") == CONFIDENCE_COLORS["high"].lower()
    assert fill_by_value.get("Drug A") == CONFIDENCE_COLORS["verified"].lower()
    assert fill_by_value.get("45") == CONFIDENCE_COLORS["medium"].lower()
    assert fill_by_value.get("Placebo") == CONFIDENCE_COLORS["low"].lower()


def test_export_with_evidence(tmp_path: Path) -> None:
    """include_evidence=True should add evidence columns next to each field."""
    output = tmp_path / "evidence.xlsx"
    export_extraction_results(
        results=_make_results(),
        field_names=FIELD_NAMES,
        output_path=output,
        include_evidence=True,
        include_confidence=False,
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Extraction Results"]

    # With evidence: col 1 = Sample Size, col 2 = Sample Size [evidence],
    #                col 3 = Intervention, col 4 = Intervention [evidence]
    assert ws.cell(1, 1).value == "Sample Size"
    assert ws.cell(1, 2).value == "Sample Size [evidence]"
    assert ws.cell(1, 3).value == "Intervention"
    assert ws.cell(1, 4).value == "Intervention [evidence]"

    # Spot-check evidence in a data row
    row2_values = {ws.cell(2, c).value for c in range(1, 5)}
    assert '{"page": 3, "text": "120 patients"}' in row2_values or \
        '{"page": 1, "text": "Drug A was given"}' in row2_values


def test_export_metadata_sheet(tmp_path: Path) -> None:
    """_MetaScreener_Log sheet must exist with export date and PDF count."""
    output = tmp_path / "meta.xlsx"
    export_extraction_results(
        results=_make_results(),
        field_names=FIELD_NAMES,
        output_path=output,
    )

    wb = openpyxl.load_workbook(output)
    assert "_MetaScreener_Log" in wb.sheetnames

    meta = wb["_MetaScreener_Log"]
    assert meta["A1"].value == "Export Date"
    assert meta["B1"].value is not None  # ISO timestamp
    assert meta["A2"].value == "Total PDFs"
    assert meta["B2"].value == 2  # two distinct pdf_ids


def test_export_empty_results(tmp_path: Path) -> None:
    """Empty results list should produce a file with only headers."""
    output = tmp_path / "empty.xlsx"
    export_extraction_results(
        results=[],
        field_names=FIELD_NAMES,
        output_path=output,
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Extraction Results"]
    # Only header row; no data rows
    assert ws.max_row == 1
    assert ws.cell(1, 1).value == "Sample Size"


def test_export_missing_field_value(tmp_path: Path) -> None:
    """PDFs missing some fields should export empty string for those cells."""
    results = [
        {"pdf_id": "pdf_X", "field_name": "Sample Size", "value": "88",
         "confidence": "high", "evidence_json": ""},
        # No Intervention field for pdf_X
    ]
    output = tmp_path / "missing.xlsx"
    export_extraction_results(
        results=results,
        field_names=FIELD_NAMES,
        output_path=output,
        include_confidence=False,
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Extraction Results"]
    assert ws.cell(2, 1).value == "88"
    # openpyxl reads empty-string writes back as None (Excel has no empty-string concept)
    assert ws.cell(2, 2).value in ("", None)  # missing field → empty / None
