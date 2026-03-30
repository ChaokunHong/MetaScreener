"""Tests for template_filler — export by filling original Excel template."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from metascreener.core.enums import FieldRole, SheetCardinality, SheetRole
from metascreener.core.models_extraction import (
    ExtractionSchema,
    FieldSchema,
    SheetSchema,
)
from metascreener.module2_extraction.export.template_filler import (
    export_filled_template,
)


def _make_template(
    path: Path,
    sheets: dict[str, list[str]],
    formulas: dict[str, dict[int, str]] | None = None,
) -> Path:
    """Create a minimal Excel template for testing.

    Args:
        path: Where to write the workbook.
        sheets: Mapping of sheet_name -> list of header names.
        formulas: Optional sheet_name -> {col_index: formula} for row 2.

    Returns:
        The path that was written.
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, headers in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(1, col_idx, header)
        if formulas and sheet_name in formulas:
            for col_idx, formula in formulas[sheet_name].items():
                ws.cell(2, col_idx, formula)
    wb.save(path)
    return path


def _cell(
    sheet: str, row: int, field: str, value: str,
) -> dict:
    """Build a result cell dict (shorthand to avoid long lines)."""
    return {
        "sheet_name": sheet,
        "row_index": row,
        "field_name": field,
        "value": value,
    }


def _make_schema(sheets: list[dict]) -> ExtractionSchema:
    """Build a minimal ExtractionSchema from dicts."""
    sheet_schemas = []
    for s in sheets:
        fields = []
        for f in s["fields"]:
            fields.append(
                FieldSchema(
                    column=f["name"][0],
                    name=f["name"],
                    description="",
                    field_type="text",
                    role=FieldRole(f["role"]),
                )
            )
        sheet_schemas.append(
            SheetSchema(
                sheet_name=s["sheet_name"],
                role=SheetRole(s.get("role", "data")),
                cardinality=SheetCardinality(
                    s.get("cardinality", "one_per_study"),
                ),
                fields=fields,
                extraction_order=s.get("extraction_order", 0),
            )
        )
    return ExtractionSchema(
        schema_id="test-schema",
        schema_version="1.0",
        sheets=sheet_schemas,
    )


class TestFillsExtractColumnsOnly:
    """Verify that only EXTRACT-role columns are written to."""

    def test_extract_columns_filled(self, tmp_path: Path) -> None:
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["ID", "Title", "CalcCol"]},
        )
        schema = _make_schema([{
            "sheet_name": "Studies",
            "fields": [
                {"name": "ID", "role": "extract"},
                {"name": "Title", "role": "extract"},
                {"name": "CalcCol", "role": "auto_calc"},
            ],
        }])
        results = [
            _cell("Studies", 0, "ID", "S001"),
            _cell("Studies", 0, "Title", "My Study"),
            _cell("Studies", 0, "CalcCol", "SHOULD_NOT_APPEAR"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        ws = wb["Studies"]
        assert ws.cell(2, 1).value == "S001"
        assert ws.cell(2, 2).value == "My Study"
        # CalcCol is auto_calc — should NOT be filled
        assert ws.cell(2, 3).value is None

    def test_auto_calc_column_untouched(self, tmp_path: Path) -> None:
        """Auto-calc columns must not be written."""
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["Score", "Grade"]},
        )
        schema = _make_schema([{
            "sheet_name": "Studies",
            "fields": [
                {"name": "Score", "role": "extract"},
                {"name": "Grade", "role": "auto_calc"},
            ],
        }])
        results = [
            _cell("Studies", 0, "Score", "95"),
            _cell("Studies", 0, "Grade", "A"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        ws = wb["Studies"]
        assert ws.cell(2, 1).value == "95"
        assert ws.cell(2, 2).value is None


class TestPreservesMappingSheets:
    """Mapping and documentation sheets must not be modified."""

    def test_mapping_sheet_unchanged(self, tmp_path: Path) -> None:
        template = _make_template(
            tmp_path / "template.xlsx",
            {
                "Studies": ["ID", "Title"],
                "Reference_Lists": ["Code", "Label"],
            },
        )
        # Write some reference data
        wb = openpyxl.load_workbook(template)
        ws = wb["Reference_Lists"]
        ws.cell(2, 1, "REF01")
        ws.cell(2, 2, "Reference Label")
        wb.save(template)

        schema = _make_schema([
            {
                "sheet_name": "Studies",
                "fields": [
                    {"name": "ID", "role": "extract"},
                    {"name": "Title", "role": "extract"},
                ],
            },
            {
                "sheet_name": "Reference_Lists",
                "role": "mapping",
                "fields": [
                    {"name": "Code", "role": "metadata"},
                    {"name": "Label", "role": "metadata"},
                ],
            },
        ])
        results = [
            _cell("Studies", 0, "ID", "S001"),
            _cell("Studies", 0, "Title", "Test"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        ws = wb["Reference_Lists"]
        assert ws.cell(2, 1).value == "REF01"
        assert ws.cell(2, 2).value == "Reference Label"


class TestManyPerStudyMultipleRows:
    """Many-per-study sheets should fill multiple rows."""

    def test_fills_multiple_rows(self, tmp_path: Path) -> None:
        sn = "Resistance_Data"
        template = _make_template(
            tmp_path / "template.xlsx",
            {sn: ["Pathogen", "Drug", "MIC"]},
        )
        schema = _make_schema([{
            "sheet_name": sn,
            "cardinality": "many_per_study",
            "fields": [
                {"name": "Pathogen", "role": "extract"},
                {"name": "Drug", "role": "extract"},
                {"name": "MIC", "role": "extract"},
            ],
        }])
        results = [
            _cell(sn, 0, "Pathogen", "E. coli"),
            _cell(sn, 0, "Drug", "Ampicillin"),
            _cell(sn, 0, "MIC", "32"),
            _cell(sn, 1, "Pathogen", "K. pneumoniae"),
            _cell(sn, 1, "Drug", "Meropenem"),
            _cell(sn, 1, "MIC", "0.5"),
            _cell(sn, 2, "Pathogen", "S. aureus"),
            _cell(sn, 2, "Drug", "Vancomycin"),
            _cell(sn, 2, "MIC", "1"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        ws = wb[sn]
        # Row 2 = row_index 0
        assert ws.cell(2, 1).value == "E. coli"
        assert ws.cell(2, 2).value == "Ampicillin"
        assert ws.cell(2, 3).value == "32"
        # Row 3 = row_index 1
        assert ws.cell(3, 1).value == "K. pneumoniae"
        assert ws.cell(3, 2).value == "Meropenem"
        assert ws.cell(3, 3).value == "0.5"
        # Row 4 = row_index 2
        assert ws.cell(4, 1).value == "S. aureus"
        assert ws.cell(4, 2).value == "Vancomycin"
        assert ws.cell(4, 3).value == "1"


class TestSkipsFormulaCells:
    """Cells containing formulas must never be overwritten."""

    def test_formula_cell_not_overwritten(
        self, tmp_path: Path,
    ) -> None:
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["Score", "Rank"]},
            formulas={"Studies": {2: '=IF(A2>50,"High","Low")'}},
        )
        schema = _make_schema([{
            "sheet_name": "Studies",
            "fields": [
                {"name": "Score", "role": "extract"},
                {"name": "Rank", "role": "extract"},
            ],
        }])
        results = [
            _cell("Studies", 0, "Score", "75"),
            _cell("Studies", 0, "Rank", "SHOULD_NOT_APPEAR"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        ws = wb["Studies"]
        assert ws.cell(2, 1).value == "75"
        # Column 2 had a formula — must be preserved
        assert str(ws.cell(2, 2).value).startswith("=")


class TestEdgeCases:
    """Edge cases: missing template, empty results, missing sheets."""

    def test_missing_template_raises(self, tmp_path: Path) -> None:
        schema = _make_schema([])
        with pytest.raises(FileNotFoundError):
            export_filled_template(
                tmp_path / "nonexistent.xlsx",
                [],
                schema,
                tmp_path / "out.xlsx",
            )

    def test_empty_results(self, tmp_path: Path) -> None:
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["ID", "Title"]},
        )
        schema = _make_schema([{
            "sheet_name": "Studies",
            "fields": [
                {"name": "ID", "role": "extract"},
                {"name": "Title", "role": "extract"},
            ],
        }])

        output = tmp_path / "filled.xlsx"
        result = export_filled_template(template, [], schema, output)
        assert result == output
        assert output.exists()

        wb = openpyxl.load_workbook(output)
        ws = wb["Studies"]
        assert ws.cell(2, 1).value is None
        assert ws.cell(2, 2).value is None

    def test_schema_sheet_not_in_workbook(
        self, tmp_path: Path,
    ) -> None:
        """Schema references a sheet not in the workbook."""
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["ID"]},
        )
        schema = _make_schema([{
            "sheet_name": "NonExistent",
            "fields": [{"name": "ID", "role": "extract"}],
        }])
        results = [
            _cell("NonExistent", 0, "ID", "X"),
        ]

        output = tmp_path / "filled.xlsx"
        # Should not raise — just skip the missing sheet
        export_filled_template(template, results, schema, output)
        assert output.exists()

    def test_non_integer_row_index(self, tmp_path: Path) -> None:
        """row_index as string should be coerced to int."""
        template = _make_template(
            tmp_path / "template.xlsx",
            {"Studies": ["ID"]},
        )
        schema = _make_schema([{
            "sheet_name": "Studies",
            "fields": [{"name": "ID", "role": "extract"}],
        }])
        results = [
            {
                "sheet_name": "Studies",
                "row_index": "0",
                "field_name": "ID",
                "value": "S001",
            },
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        assert wb["Studies"].cell(2, 1).value == "S001"


class TestMultipleDataSheets:
    """Verify filling across multiple data sheets."""

    def test_fills_two_data_sheets(self, tmp_path: Path) -> None:
        template = _make_template(
            tmp_path / "template.xlsx",
            {
                "Study_Characteristics": ["Author", "Year"],
                "Pathogen_Summary": ["Pathogen", "Count"],
                "Data_Dictionary": ["Term", "Definition"],
            },
        )
        schema = _make_schema([
            {
                "sheet_name": "Study_Characteristics",
                "extraction_order": 0,
                "fields": [
                    {"name": "Author", "role": "extract"},
                    {"name": "Year", "role": "extract"},
                ],
            },
            {
                "sheet_name": "Pathogen_Summary",
                "extraction_order": 1,
                "fields": [
                    {"name": "Pathogen", "role": "extract"},
                    {"name": "Count", "role": "extract"},
                ],
            },
            {
                "sheet_name": "Data_Dictionary",
                "role": "documentation",
                "extraction_order": 99,
                "fields": [
                    {"name": "Term", "role": "metadata"},
                    {"name": "Definition", "role": "metadata"},
                ],
            },
        ])
        sc = "Study_Characteristics"
        ps = "Pathogen_Summary"
        results = [
            _cell(sc, 0, "Author", "Smith"),
            _cell(sc, 0, "Year", "2024"),
            _cell(ps, 0, "Pathogen", "E. coli"),
            _cell(ps, 0, "Count", "42"),
        ]

        output = tmp_path / "filled.xlsx"
        export_filled_template(template, results, schema, output)

        wb = openpyxl.load_workbook(output)
        assert wb[sc].cell(2, 1).value == "Smith"
        assert wb[sc].cell(2, 2).value == "2024"
        assert wb[ps].cell(2, 1).value == "E. coli"
        assert wb[ps].cell(2, 2).value == "42"
        # Documentation sheet should be untouched
        assert wb["Data_Dictionary"].cell(2, 1).value is None
