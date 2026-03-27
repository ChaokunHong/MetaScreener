"""Tests for Excel exporter."""
from __future__ import annotations
from pathlib import Path
import openpyxl
import pytest
from metascreener.core.enums import Confidence, FieldRole, SheetCardinality, SheetRole
from metascreener.core.models_extraction import (
    CellValue, ExtractionSchema, ExtractionSessionResult, FieldSchema, RowResult, SheetResult, SheetSchema,
)
from metascreener.module2_extraction.exporter import export_to_excel

def _make_schema() -> ExtractionSchema:
    return ExtractionSchema(
        schema_id="test", schema_version="1.0",
        sheets=[SheetSchema(
            sheet_name="Study", role=SheetRole.DATA, cardinality=SheetCardinality.ONE_PER_STUDY,
            fields=[
                FieldSchema(column="A", name="row_id", description="ID", field_type="number", role=FieldRole.AUTO_CALC),
                FieldSchema(column="B", name="author", description="Author", field_type="text", role=FieldRole.EXTRACT, required=True),
                FieldSchema(column="C", name="year", description="Year", field_type="number", role=FieldRole.EXTRACT),
            ], extraction_order=1,
        )],
    )

def _make_results() -> list[ExtractionSessionResult]:
    return [
        ExtractionSessionResult(pdf_id="p1", pdf_filename="Smith_2023.pdf", sheets={
            "Study": SheetResult(sheet_name="Study", rows=[
                RowResult(row_index=0, fields={
                    "author": CellValue(value="Smith", confidence=Confidence.HIGH),
                    "year": CellValue(value=2023, confidence=Confidence.HIGH),
                }),
            ]),
        }),
        ExtractionSessionResult(pdf_id="p2", pdf_filename="Jones_2024.pdf", sheets={
            "Study": SheetResult(sheet_name="Study", rows=[
                RowResult(row_index=0, fields={
                    "author": CellValue(value="Jones", confidence=Confidence.MEDIUM),
                    "year": CellValue(value=2024, confidence=Confidence.LOW),
                }),
            ]),
        }),
    ]

class TestExportToExcel:
    def test_creates_file(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        export_to_excel(schema=_make_schema(), results=_make_results(), output_path=out)
        assert out.exists()

    def test_has_data_sheets(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        export_to_excel(schema=_make_schema(), results=_make_results(), output_path=out)
        wb = openpyxl.load_workbook(out)
        assert "Study" in wb.sheetnames

    def test_writes_correct_values(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        export_to_excel(schema=_make_schema(), results=_make_results(), output_path=out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Study"]
        assert ws.cell(row=1, column=1).value == "author"
        assert ws.cell(row=1, column=2).value == "year"
        assert ws.cell(row=2, column=1).value == "Smith"
        assert ws.cell(row=2, column=2).value == 2023
        assert ws.cell(row=3, column=1).value == "Jones"

    def test_skips_auto_calc_fields(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        export_to_excel(schema=_make_schema(), results=_make_results(), output_path=out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Study"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10) if ws.cell(row=1, column=c).value]
        assert "row_id" not in headers

    def test_has_log_sheet(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        export_to_excel(schema=_make_schema(), results=_make_results(), output_path=out)
        wb = openpyxl.load_workbook(out)
        assert "_MetaScreener_Log" in wb.sheetnames
